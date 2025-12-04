from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify
from flask_bcrypt import Bcrypt
import psycopg2
from psycopg2.extras import RealDictCursor
from psycopg2.extensions import ISOLATION_LEVEL_AUTOCOMMIT
from dotenv import load_dotenv
import os
import hashlib
from datetime import datetime, date, timedelta
from functools import wraps
import io
import csv
import subprocess
import json
import threading
import schedule
import time
from pathlib import Path
import pandas as pd
from werkzeug.middleware.proxy_fix import ProxyFix

# Load environment variables
load_dotenv()

# Initialize Flask app
app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)
app.secret_key = os.getenv('SECRET_KEY', 'dev-secret-key-change-in-production')
bcrypt = Bcrypt(app)

# Security headers
@app.after_request
def security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response

# Database configuration
DB_CONFIG = {
    'dbname': os.getenv('DB_NAME', 'raman_research_prod'),
    'user': os.getenv('DB_USER', 'postgres'),
    'password': os.getenv('DB_PASSWORD'),
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': os.getenv('DB_PORT', '5432')
}

# Configuration for starting Patient ID
STARTING_PATIENT_ID = int(os.getenv('STARTING_PATIENT_ID', '1500'))

# Backup configuration
BACKUP_CONFIG_FILE = os.getenv('BACKUP_CONFIG_FILE', 'backup_config.json')
DEFAULT_BACKUP_DIR = os.getenv('BACKUP_DIRECTORY', '/backups')
DEFAULT_RETENTION_DAYS = int(os.getenv('BACKUP_RETENTION_DAYS', '90'))

# Global scheduler variables
scheduler_thread = None
scheduler_running = False


def create_database_if_not_exists():
    """Create the database if it doesn't exist"""
    try:
        print(f"Checking if database '{DB_CONFIG['dbname']}' exists...")
        print(f"Connecting to PostgreSQL server at {DB_CONFIG['host']}:{DB_CONFIG['port']}")
        print(f"Using user: {DB_CONFIG['user']}")

        # Connect to postgres database to check/create the database
        conn = psycopg2.connect(
            dbname='postgres',
            user=DB_CONFIG['user'],
            password=DB_CONFIG['password'],
            host=DB_CONFIG['host'],
            port=DB_CONFIG['port']
        )
        conn.set_isolation_level(ISOLATION_LEVEL_AUTOCOMMIT)
        cur = conn.cursor()

        # Check if database exists
        cur.execute("SELECT 1 FROM pg_database WHERE datname = %s", (DB_CONFIG['dbname'],))
        exists = cur.fetchone()

        if not exists:
            print(f"Database '{DB_CONFIG['dbname']}' does not exist. Creating...")
            cur.execute(f"CREATE DATABASE {DB_CONFIG['dbname']}")
            print(f"✓ Database '{DB_CONFIG['dbname']}' created successfully")
        else:
            print(f"✓ Database '{DB_CONFIG['dbname']}' already exists")

        cur.close()
        conn.close()
        return True

    except psycopg2.Error as e:
        print(f"✗ Error with database: {e}")
        print(f"\nPlease ensure PostgreSQL is running and create the database manually:")
        print(f"  CREATE DATABASE {DB_CONFIG['dbname']};")
        return False


def get_db_connection():
    """Create and return a database connection"""
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        return conn
    except Exception as e:
        print(f"Database connection error: {e}")
        return None


def init_database():
    """Initialize database with all required tables and ICD-10 codes from Excel if available"""
    conn = get_db_connection()
    if not conn:
        print("Failed to connect to database")
        return False

    try:
        cur = conn.cursor()

        print("Configuring tables...")

        # Create users table
        cur.execute('''
            CREATE TABLE IF NOT EXISTS users (
                user_id SERIAL PRIMARY KEY,
                username VARCHAR(50) UNIQUE NOT NULL,
                password_hash VARCHAR(255) NOT NULL,
                email VARCHAR(100),
                role VARCHAR(20) NOT NULL CHECK (role IN ('Administrator', 'Staff', 'Patient')),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_login TIMESTAMP
            )
        ''')

        # Create patients_sensitive table (PII data)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS patients_sensitive (
                patient_id INTEGER PRIMARY KEY CHECK (patient_id >= 1 AND patient_id <= 99999),
                patient_name VARCHAR(255) NOT NULL,
                mbo VARCHAR(9) NOT NULL,
                date_of_birth DATE,
                date_of_sample_collection DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create patients_statistical table (anonymized data)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS patients_statistical (
                patient_id INTEGER PRIMARY KEY REFERENCES patients_sensitive(patient_id) ON DELETE CASCADE,
                person_hash VARCHAR(64) NOT NULL,
                age INTEGER,
                sex VARCHAR(1) CHECK (sex IN ('M', 'F')),
                eye VARCHAR(3) CHECK (eye IN ('L', 'R', 'ND')),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create ocular_conditions table (main conditions - one row per patient)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS ocular_conditions (
                patient_id INTEGER PRIMARY KEY REFERENCES patients_sensitive(patient_id) ON DELETE CASCADE,
                lens_status VARCHAR(20) CHECK (lens_status IN ('Phakic', 'Pseudophakic', 'Aphakic', 'ND')),
                locs_iii_no VARCHAR(10),
                locs_iii_nc VARCHAR(10),
                locs_iii_c VARCHAR(10),
                locs_iii_p VARCHAR(10),
                iol_type VARCHAR(50),
                etiology_aphakia VARCHAR(50),
                glaucoma VARCHAR(10),
                oht_or_pac VARCHAR(10),
                etiology_glaucoma VARCHAR(50),
                steroid_responder VARCHAR(10),
                pxs VARCHAR(10),
                pds VARCHAR(10),
                diabetic_retinopathy VARCHAR(10),
                stage_diabetic_retinopathy VARCHAR(50),
                stage_npdr VARCHAR(50),
                stage_pdr VARCHAR(50),
                macular_edema VARCHAR(10),
                etiology_macular_edema VARCHAR(50),
                macular_degeneration_dystrophy VARCHAR(10),
                etiology_macular_deg_dyst VARCHAR(50),
                stage_amd VARCHAR(50),
                exudation_amd VARCHAR(10),
                stage_other_macular_deg VARCHAR(50),
                exudation_other_macular_deg VARCHAR(10),
                macular_hole_vmt VARCHAR(10),
                etiology_mh_vmt VARCHAR(50),
                cause_secondary_mh_vmt TEXT,
                treatment_status_mh_vmt VARCHAR(50),
                epiretinal_membrane VARCHAR(10),
                etiology_erm VARCHAR(50),
                cause_secondary_erm TEXT,
                treatment_status_erm VARCHAR(50),
                retinal_detachment VARCHAR(50),
                etiology_rd VARCHAR(100),
                treatment_status_rd VARCHAR(100),
                pvr VARCHAR(10),
                vitreous_haemorrhage_opacification VARCHAR(50),
                etiology_vitreous_haemorrhage VARCHAR(100),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create other_ocular_conditions table (one-to-many)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS other_ocular_conditions (
                id SERIAL PRIMARY KEY,
                patient_id INTEGER REFERENCES patients_sensitive(patient_id) ON DELETE CASCADE,
                icd10_code VARCHAR(20) NOT NULL,
                eye VARCHAR(10) CHECK (eye IN ('R', 'L', 'R+L', 'ND')),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create previous_ocular_surgeries table (one-to-many)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS previous_ocular_surgeries (
                id SERIAL PRIMARY KEY,
                patient_id INTEGER REFERENCES patients_sensitive(patient_id) ON DELETE CASCADE,
                surgery_code VARCHAR(100) NOT NULL,
                eye VARCHAR(10) CHECK (eye IN ('R', 'L', 'R+L', 'ND')),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create systemic_conditions table (one-to-many)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS systemic_conditions (
                id SERIAL PRIMARY KEY,
                patient_id INTEGER REFERENCES patients_sensitive(patient_id) ON DELETE CASCADE,
                icd10_code VARCHAR(20) NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create ocular_medications table (one-to-many)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS ocular_medications (
                id SERIAL PRIMARY KEY,
                patient_id INTEGER REFERENCES patients_sensitive(patient_id) ON DELETE CASCADE,
                trade_name VARCHAR(255) NOT NULL,
                generic_name VARCHAR(255) NOT NULL,
                eye VARCHAR(10) CHECK (eye IN ('R', 'L', 'R+L', 'ND')),
                last_application_days INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create systemic_medications table (one-to-many)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS systemic_medications (
                id SERIAL PRIMARY KEY,
                patient_id INTEGER REFERENCES patients_sensitive(patient_id) ON DELETE CASCADE,
                trade_name VARCHAR(255) NOT NULL,
                generic_name VARCHAR(255) NOT NULL,
                last_application_days INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Create reference data tables
        cur.execute('''
            CREATE TABLE IF NOT EXISTS icd10_ocular_conditions (
                id SERIAL PRIMARY KEY,
                code VARCHAR(20) UNIQUE NOT NULL,
                description TEXT NOT NULL,
                category VARCHAR(100),
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cur.execute('''
            CREATE TABLE IF NOT EXISTS icd10_systemic_conditions (
                id SERIAL PRIMARY KEY,
                code VARCHAR(20) UNIQUE NOT NULL,
                description TEXT NOT NULL,
                category VARCHAR(100),
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cur.execute('''
            CREATE TABLE IF NOT EXISTS medications (
                id SERIAL PRIMARY KEY,
                trade_name VARCHAR(255) NOT NULL,
                generic_name VARCHAR(255) NOT NULL,
                medication_type VARCHAR(20) CHECK (medication_type IN ('Ocular', 'Systemic', 'Both')),
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        cur.execute('''
            CREATE TABLE IF NOT EXISTS surgeries (
                id SERIAL PRIMARY KEY,
                code VARCHAR(100) UNIQUE NOT NULL,
                description TEXT NOT NULL,
                category VARCHAR(100),
                active BOOLEAN DEFAULT TRUE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        conn.commit()
        print("✓ Tables configured successfully")

        cur.close()
        conn.close()

        # Populate reference data in tables
        populate_reference_data()

        return True

    except Exception as e:
        print(f"✗ Error initializing database: {e}")
        if conn:
            conn.rollback()
            conn.close()
        return False


def populate_reference_data():
    """Populate reference data tables if they're empty"""
    conn = get_db_connection()
    if not conn:
        print("Failed to connect to database for data population")
        return False

    try:
        cur = conn.cursor()

        # Check and populate default admin user
        cur.execute("SELECT COUNT(*) FROM users")
        if cur.fetchone()[0] == 0:
            admin_password = bcrypt.generate_password_hash('admin123').decode('utf-8')
            cur.execute('''
                INSERT INTO users (username, password_hash, email, role)
                VALUES (%s, %s, %s, %s)
            ''', ('Admin', admin_password, '', 'Administrator'))
            conn.commit()
            print("✓ Default admin user created (username: Admin, password: admin123)")
            print("  ⚠️  IMPORTANT: Change the admin password after first login!")

        # Populate ICD-10 codes
        populate_icd10_codes(conn, cur)

        # Populate medications
        populate_medications(conn, cur)

        # Populate surgeries
        populate_surgeries(conn, cur)

        cur.close()
        conn.close()
        return True

    except Exception as e:
        print(f"Error populating reference data: {e}")
        if conn:
            conn.rollback()
            conn.close()
        return False


def populate_icd10_codes(conn, cur):
    """Populate ICD-10 codes from Excel files or defaults"""

    # Check and populate ocular codes
    cur.execute("SELECT COUNT(*) FROM icd10_ocular_conditions")
    ocular_count = cur.fetchone()[0]  # Store the count in a variable!

    if ocular_count == 0:
        print("ICD-10 ocular table is empty, populating...")
        ocular_file = 'ICD10_eye_codes.xlsx'

        if os.path.exists(ocular_file):
            try:
                import pandas as pd
                print(f"Importing ICD-10 ocular codes from {ocular_file}...")

                df = pd.read_excel(ocular_file)
                imported_count = 0

                for _, row in df.iterrows():
                    code = str(row['ICD-10 Code']).strip()
                    description = str(row['Description']).strip()

                    # Determine category based on code prefix
                    category = get_ocular_category(code)

                    try:
                        cur.execute('''
                            INSERT INTO icd10_ocular_conditions (code, description, category, active)
                            VALUES (%s, %s, %s, TRUE)
                            ON CONFLICT (code) DO NOTHING
                        ''', (code, description, category))
                        imported_count += 1
                    except Exception as e:
                        print(f"  Skipped code {code}: {e}")

                conn.commit()
                print(f"✓ Imported {imported_count} ICD-10 ocular codes from Excel")

            except ImportError as e:
                print(f"⚠ pandas not installed or error: {e}")
                insert_default_ocular_codes(conn, cur)
            except Exception as e:
                print(f"⚠ Error importing from Excel: {e}")
                insert_default_ocular_codes(conn, cur)
        else:
            print(f"Excel file {ocular_file} not found")
            insert_default_ocular_codes(conn, cur)
    else:
        print(f"ICD-10 ocular table has data ({ocular_count} codes), skipping population")  # Use the stored variable!

    # Check and populate systemic codes
    cur.execute("SELECT COUNT(*) FROM icd10_systemic_conditions")
    systemic_count = cur.fetchone()[0]  # Store the count in a variable!

    if systemic_count == 0:
        print("ICD-10 systemic table is empty, populating...")
        systemic_file = 'ICD10_non_eye_codes.xlsx'

        if os.path.exists(systemic_file):
            try:
                import pandas as pd
                print(f"Importing ICD-10 systemic codes from {systemic_file}...")

                df = pd.read_excel(systemic_file)
                imported_count = 0

                for _, row in df.iterrows():
                    code = str(row['ICD-10 Code']).strip()
                    description = str(row['Description']).strip()

                    # Determine category based on first letter
                    category = get_systemic_category(code)

                    try:
                        cur.execute('''
                            INSERT INTO icd10_systemic_conditions (code, description, category, active)
                            VALUES (%s, %s, %s, TRUE)
                            ON CONFLICT (code) DO NOTHING
                        ''', (code, description, category))
                        imported_count += 1
                    except Exception as e:
                        print(f"  Skipped code {code}: {e}")

                conn.commit()
                print(f"✓ Imported {imported_count} ICD-10 systemic codes from Excel")

            except ImportError as e:
                print(f"⚠ pandas not installed or error: {e}")
                insert_default_systemic_codes(conn, cur)
            except Exception as e:
                print(f"⚠ Error importing from Excel: {e}")
                insert_default_systemic_codes(conn, cur)
        else:
            print(f"Excel file {systemic_file} not found")
            insert_default_systemic_codes(conn, cur)
    else:
        print(f"ICD-10 systemic table has data ({systemic_count} codes), skipping population")


def get_ocular_category(code):
    """Determine category for ocular ICD-10 code"""
    if code.startswith(('H00', 'H01', 'H02', 'H04', 'H05')):
        return 'Eyelid, lacrimal system and orbit'
    elif code.startswith(('H10', 'H11', 'H15', 'H16', 'H17', 'H18')):
        return 'Conjunctiva, sclera and cornea'
    elif code.startswith(('H20', 'H21', 'H22')):
        return 'Iris and ciliary body'
    elif code.startswith(('H25', 'H26', 'H27', 'H28')):
        return 'Lens disorders'
    elif code.startswith(('H30', 'H31', 'H32', 'H33', 'H34', 'H35', 'H36')):
        return 'Choroid and retina'
    elif code.startswith(('H40', 'H42')):
        return 'Glaucoma'
    elif code.startswith(('H43', 'H44')):
        return 'Vitreous and globe'
    elif code.startswith(('H46', 'H47', 'H48')):
        return 'Optic nerve and visual pathways'
    elif code.startswith(('H49', 'H50', 'H51', 'H52')):
        return 'Ocular muscles and refraction'
    elif code.startswith(('H53', 'H54')):
        return 'Visual disturbances and blindness'
    elif code.startswith(('H55', 'H57', 'H59')):
        return 'Other disorders of eye'
    return None


def get_systemic_category(code):
    """Determine category for systemic ICD-10 code"""
    if not code:
        return None

    first_char = code[0]

    if first_char in ('A', 'B'):
        return 'Infectious diseases'
    elif first_char == 'C':
        return 'Neoplasms - malignant'
    elif first_char == 'D':
        if len(code) > 2 and code[1:3].isdigit() and int(code[1:3]) < 50:
            return 'Neoplasms - benign/other'
        else:
            return 'Blood and immune disorders'
    elif first_char == 'E':
        return 'Endocrine and metabolic'
    elif first_char == 'F':
        return 'Mental and behavioral'
    elif first_char == 'G':
        return 'Nervous system'
    elif first_char == 'I':
        return 'Circulatory system'
    elif first_char == 'J':
        return 'Respiratory system'
    elif first_char == 'K':
        return 'Digestive system'
    elif first_char == 'L':
        return 'Skin and subcutaneous'
    elif first_char == 'M':
        return 'Musculoskeletal'
    elif first_char == 'N':
        return 'Genitourinary'
    elif first_char == 'O':
        return 'Pregnancy and childbirth'
    elif first_char == 'P':
        return 'Perinatal conditions'
    elif first_char == 'Q':
        return 'Congenital malformations'
    elif first_char == 'R':
        return 'Symptoms and signs'
    elif first_char in ('S', 'T'):
        return 'Injury and poisoning'
    elif first_char in ('V', 'W', 'X', 'Y'):
        return 'External causes'
    elif first_char == 'Z':
        return 'Health status factors'

    return 'Other'


def insert_default_ocular_codes(conn, cur):
    """Insert default ocular ICD-10 codes"""
    print("Inserting default ocular codes...")
    ocular_codes = [
        ('H25.9', 'Senile cataract, unspecified', 'Lens disorders'),
        ('H26.9', 'Cataract, unspecified', 'Lens disorders'),
        ('H40.1', 'Primary open-angle glaucoma', 'Glaucoma'),
        ('H40.2', 'Primary angle-closure glaucoma', 'Glaucoma'),
        ('H35.3', 'Degeneration of macula and posterior pole', 'Choroid and retina'),
        ('H36.0', 'Diabetic retinopathy', 'Choroid and retina'),
        ('H33.0', 'Retinal detachment with retinal break', 'Choroid and retina'),
        ('H34.8', 'Other retinal vascular occlusions', 'Choroid and retina'),
        ('H10.1', 'Acute atopic conjunctivitis', 'Conjunctiva, sclera and cornea'),
        ('H16.0', 'Corneal ulcer', 'Conjunctiva, sclera and cornea'),
    ]

    for code, description, category in ocular_codes:
        try:
            cur.execute('''
                INSERT INTO icd10_ocular_conditions (code, description, category, active)
                VALUES (%s, %s, %s, TRUE)
                ON CONFLICT (code) DO NOTHING
            ''', (code, description, category))
        except Exception as e:
            print(f"  Error inserting {code}: {e}")

    conn.commit()
    print(f"✓ Inserted {len(ocular_codes)} default ICD-10 ocular codes")


def insert_default_systemic_codes(conn, cur):
    """Insert default systemic ICD-10 codes"""
    print("Inserting default systemic codes...")
    systemic_codes = [
        ('E11.9', 'Type 2 diabetes mellitus without complications', 'Endocrine and metabolic'),
        ('I10', 'Essential (primary) hypertension', 'Circulatory system'),
        ('E78.5', 'Hyperlipidemia, unspecified', 'Endocrine and metabolic'),
        ('J45.9', 'Asthma, unspecified', 'Respiratory system'),
        ('M79.3', 'Myalgia', 'Musculoskeletal'),
        ('F32.9', 'Major depressive disorder, single episode, unspecified', 'Mental and behavioral'),
        ('G43.9', 'Migraine, unspecified', 'Nervous system'),
        ('K21.9', 'Gastro-esophageal reflux disease without esophagitis', 'Digestive system'),
    ]

    for code, description, category in systemic_codes:
        try:
            cur.execute('''
                INSERT INTO icd10_systemic_conditions (code, description, category, active)
                VALUES (%s, %s, %s, TRUE)
                ON CONFLICT (code) DO NOTHING
            ''', (code, description, category))
        except Exception as e:
            print(f"  Error inserting {code}: {e}")

    conn.commit()
    print(f"✓ Inserted {len(systemic_codes)} default ICD-10 systemic codes")


def populate_medications(conn, cur):
    """Populate medications table if empty"""
    cur.execute("SELECT COUNT(*) FROM medications")
    if cur.fetchone()[0] == 0:
        print("Medications table is empty, populating with defaults...")
        sample_medications = [
            # Ocular medications
            ('Timolol 0.5%', 'Timolol', 'Ocular'),
            ('Xalatan', 'Latanoprost', 'Ocular'),
            ('Cosopt', 'Dorzolamide/Timolol', 'Ocular'),
            ('Lumigan', 'Bimatoprost', 'Ocular'),
            ('Pred Forte', 'Prednisolone acetate 1%', 'Ocular'),
            ('Vigamox', 'Moxifloxacin', 'Ocular'),
            # Systemic medications
            ('Aspirin', 'Acetylsalicylic acid', 'Systemic'),
            ('Metformin', 'Metformin', 'Systemic'),
            ('Insulin', 'Insulin', 'Systemic'),
            ('Atorvastatin', 'Atorvastatin', 'Systemic'),
            ('Amlodipine', 'Amlodipine', 'Systemic'),
        ]

        for trade, generic, med_type in sample_medications:
            try:
                cur.execute('''
                    INSERT INTO medications (trade_name, generic_name, medication_type, active)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT DO NOTHING
                ''', (trade, generic, med_type))
            except Exception as e:
                print(f"  Error inserting medication {trade}: {e}")

        conn.commit()
        print(f"✓ Inserted {len(sample_medications)} sample medications")


def populate_surgeries(conn, cur):
    """Populate surgeries table if empty"""
    cur.execute("SELECT COUNT(*) FROM surgeries")
    if cur.fetchone()[0] == 0:
        print("Surgeries table is empty, populating with defaults...")
        sample_surgeries = [
            ('PHACO', 'Phacoemulsification', 'Cataract'),
            ('ECCE', 'Extracapsular cataract extraction', 'Cataract'),
            ('IOL', 'Intraocular lens implantation', 'Cataract'),
            ('TRAB', 'Trabeculectomy', 'Glaucoma'),
            ('SLT', 'Selective laser trabeculoplasty', 'Glaucoma'),
            ('PPV', 'Pars plana vitrectomy', 'Retinal'),
            ('RD', 'Retinal detachment repair', 'Retinal'),
            ('LASIK', 'Laser-assisted in situ keratomileusis', 'Refractive'),
            ('PRK', 'Photorefractive keratectomy', 'Refractive'),
        ]

        for code, description, category in sample_surgeries:
            try:
                cur.execute('''
                    INSERT INTO surgeries (code, description, category, active)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (code) DO NOTHING
                ''', (code, description, category))
            except Exception as e:
                print(f"  Error inserting surgery {code}: {e}")

        conn.commit()
        print(f"✓ Inserted {len(sample_surgeries)} sample surgical procedures")


# Authentication and Utility Functions

def login_required(f):
    """Decorator to require login for routes"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


def admin_required(f):
    """Decorator to require administrator role"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        if session.get('role') != 'Administrator':
            flash('Administrator access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)

    return decorated_function


def staff_or_admin_required(f):
    """Decorator to require staff or administrator role"""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Please log in to access this page.', 'error')
            return redirect(url_for('login'))
        if session.get('role') not in ['Administrator', 'Staff']:
            flash('Staff or Administrator access required.', 'error')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)

    return decorated_function


def generate_person_hash(mbo):
    """Generate SHA-256 hash from MBO"""
    return hashlib.sha256(mbo.encode()).hexdigest()


def calculate_age(date_of_birth, date_of_sample):
    """Calculate age at sample collection"""
    if not date_of_birth or not date_of_sample:
        return None
    age = date_of_sample.year - date_of_birth.year
    if (date_of_sample.month, date_of_sample.day) < (date_of_birth.month, date_of_birth.day):
        age -= 1
    return age


def get_next_available_patient_id():
    """Get next available patient ID based on highest existing ID in database"""
    conn = get_db_connection()
    if not conn:
        return None
    try:
        cur = conn.cursor()
        # Get the highest patient_id currently in use
        cur.execute("SELECT COALESCE(MAX(patient_id), %s) FROM patients_sensitive", (STARTING_PATIENT_ID - 1,))
        max_id = cur.fetchone()[0]
        next_id = max_id + 1

        # Make sure we don't exceed the maximum allowed ID
        if next_id > 99999:
            cur.close()
            conn.close()
            return None

        cur.close()
        conn.close()
        return next_id
    except Exception as e:
        print(f"Error getting next patient ID: {e}")
        if conn:
            conn.close()
        return None


def check_patient_id_exists(patient_id):
    """Check if patient ID already exists"""
    conn = get_db_connection()
    if not conn:
        return False
    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM patients_sensitive WHERE patient_id = %s", (patient_id,))
        exists = cur.fetchone()[0] > 0
        cur.close()
        conn.close()
        return exists
    except Exception as e:
        print(f"Error checking patient ID: {e}")
        if conn:
            conn.close()
        return False


def build_filter_clause(request_form):
    """
    Build WHERE clause and parameters for filtering patients based on form data

    Returns:
        tuple: (where_clause, params_list)
            where_clause: SQL WHERE conditions as string
            params_list: List of parameter values for parameterized query
    """
    where_clauses = []
    params = []

    # ============================================================================
    # MAIN OCULAR CONDITIONS FILTERS
    # ============================================================================

    # Glaucoma filter
    glaucoma_filter = request_form.get('filter_glaucoma', '')
    if glaucoma_filter:
        if glaucoma_filter == 'all':
            # Include all patients (no filter needed, just documenting)
            pass
        elif glaucoma_filter == '0':
            where_clauses.append('oc.glaucoma = %s')
            params.append('0')  # FIXED: String instead of integer
        elif glaucoma_filter == '1':
            where_clauses.append('oc.glaucoma = %s')
            params.append('1')  # FIXED: String instead of integer
        elif glaucoma_filter == 'ND':
            where_clauses.append("(oc.glaucoma IS NULL OR oc.glaucoma = 'ND')")
        elif glaucoma_filter == 'not_0_not_nd':
            where_clauses.append("(oc.glaucoma IS NOT NULL AND oc.glaucoma != 'ND' AND oc.glaucoma != '0')")

    # Diabetic Retinopathy filter
    dr_filter = request_form.get('filter_diabetic_retinopathy', '')
    if dr_filter:
        if dr_filter == 'all':
            pass
        elif dr_filter == '0':
            where_clauses.append('oc.diabetic_retinopathy = %s')
            params.append('0')  # FIXED: String instead of integer
        elif dr_filter == '1':
            where_clauses.append('oc.diabetic_retinopathy = %s')
            params.append('1')  # FIXED: String instead of integer
        elif dr_filter == 'ND':
            where_clauses.append("(oc.diabetic_retinopathy IS NULL OR oc.diabetic_retinopathy = 'ND')")
        elif dr_filter == 'not_0_not_nd':
            where_clauses.append(
                "(oc.diabetic_retinopathy IS NOT NULL AND oc.diabetic_retinopathy != 'ND' AND oc.diabetic_retinopathy != '0')")

    # Lens Status filter
    lens_filter = request_form.get('filter_lens_status', '')
    if lens_filter:
        if lens_filter == 'all':
            pass
        elif lens_filter == 'ND':
            where_clauses.append("(oc.lens_status IS NULL OR oc.lens_status = 'ND')")
        elif lens_filter in ['Phakic', 'Pseudophakic', 'Aphakic']:
            where_clauses.append('oc.lens_status = %s')
            params.append(lens_filter)

    # Macular Edema filter
    me_filter = request_form.get('filter_macular_edema', '')
    if me_filter:
        if me_filter == 'all':
            pass
        elif me_filter == '0':
            where_clauses.append('oc.macular_edema = %s')
            params.append('0')  # FIXED: String instead of integer
        elif me_filter == '1':
            where_clauses.append('oc.macular_edema = %s')
            params.append('1')  # FIXED: String instead of integer
        elif me_filter == 'ND':
            where_clauses.append("(oc.macular_edema IS NULL OR oc.macular_edema = 'ND')")
        elif me_filter == 'not_0_not_nd':
            where_clauses.append(
                "(oc.macular_edema IS NOT NULL AND oc.macular_edema != 'ND' AND oc.macular_edema != '0')")

    # Macular Degeneration filter
    md_filter = request_form.get('filter_macular_degeneration', '')
    if md_filter:
        if md_filter == 'all':
            pass
        elif md_filter == '0':
            where_clauses.append('oc.macular_degeneration_dystrophy = %s')
            params.append('0')  # FIXED: String instead of integer
        elif md_filter == '1':
            where_clauses.append('oc.macular_degeneration_dystrophy = %s')
            params.append('1')  # FIXED: String instead of integer
        elif md_filter == 'ND':
            where_clauses.append(
                "(oc.macular_degeneration_dystrophy IS NULL OR oc.macular_degeneration_dystrophy = 'ND')")
        elif md_filter == 'not_0_not_nd':
            where_clauses.append(
                "(oc.macular_degeneration_dystrophy IS NOT NULL AND oc.macular_degeneration_dystrophy != 'ND' AND oc.macular_degeneration_dystrophy != '0')")

    # Epiretinal Membrane filter
    erm_filter = request_form.get('filter_epiretinal_membrane', '')
    if erm_filter:
        if erm_filter == 'all':
            pass
        elif erm_filter == '0':
            where_clauses.append('oc.epiretinal_membrane = %s')
            params.append('0')  # FIXED: String instead of integer
        elif erm_filter == '1':
            where_clauses.append('oc.epiretinal_membrane = %s')
            params.append('1')  # FIXED: String instead of integer
        elif erm_filter == 'ND':
            where_clauses.append("(oc.epiretinal_membrane IS NULL OR oc.epiretinal_membrane = 'ND')")
        elif erm_filter == 'not_0_not_nd':
            where_clauses.append(
                "(oc.epiretinal_membrane IS NOT NULL AND oc.epiretinal_membrane != 'ND' AND oc.epiretinal_membrane != '0')")

    # ============================================================================
    # REPEATABLE CATEGORY FILTERS (Other Conditions, Surgeries, Medications)
    # ============================================================================

    # Other Ocular Conditions filter
    other_ocular_filter = request_form.get('filter_other_ocular_mode', '')
    if other_ocular_filter:
        if other_ocular_filter == 'all':
            # Patient has at least one other ocular condition
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM other_ocular_conditions ooc 
                    WHERE ooc.patient_id = ps.patient_id
                )
            ''')
        elif other_ocular_filter == '0':
            # Patient has no other ocular conditions
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM other_ocular_conditions ooc 
                    WHERE ooc.patient_id = ps.patient_id
                )
            ''')
        elif other_ocular_filter == 'ND':
            # Typically ND for repeatable categories means no data
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM other_ocular_conditions ooc 
                    WHERE ooc.patient_id = ps.patient_id
                )
            ''')
        elif other_ocular_filter == 'not_0_not_nd':
            # Patient has at least one condition
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM other_ocular_conditions ooc 
                    WHERE ooc.patient_id = ps.patient_id
                )
            ''')

    # Surgeries filter
    surgeries_filter = request_form.get('filter_surgeries_mode', '')
    if surgeries_filter:
        if surgeries_filter == 'all':
            # Patient has at least one surgery
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM previous_ocular_surgeries pos 
                    WHERE pos.patient_id = ps.patient_id
                )
            ''')
        elif surgeries_filter == '0':
            # Patient has no surgeries
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM previous_ocular_surgeries pos 
                    WHERE pos.patient_id = ps.patient_id
                )
            ''')
        elif surgeries_filter == 'ND':
            # No surgeries
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM previous_ocular_surgeries pos 
                    WHERE pos.patient_id = ps.patient_id
                )
            ''')
        elif surgeries_filter == 'not_0_not_nd':
            # Patient has at least one surgery
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM previous_ocular_surgeries pos 
                    WHERE pos.patient_id = ps.patient_id
                )
            ''')

    # Ocular Medications filter
    ocular_meds_filter = request_form.get('filter_ocular_meds_mode', '')
    if ocular_meds_filter:
        if ocular_meds_filter == 'all':
            # Patient has at least one ocular medication
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM ocular_medications om 
                    WHERE om.patient_id = ps.patient_id
                )
            ''')
        elif ocular_meds_filter == '0':
            # Patient has no ocular medications
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM ocular_medications om 
                    WHERE om.patient_id = ps.patient_id
                )
            ''')
        elif ocular_meds_filter == 'ND':
            # No ocular medications
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM ocular_medications om 
                    WHERE om.patient_id = ps.patient_id
                )
            ''')
        elif ocular_meds_filter == 'not_0_not_nd':
            # Patient has at least one ocular medication
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM ocular_medications om 
                    WHERE om.patient_id = ps.patient_id
                )
            ''')

    # Systemic Medications filter
    systemic_meds_filter = request_form.get('filter_systemic_meds_mode', '')
    if systemic_meds_filter:
        if systemic_meds_filter == 'all':
            # Patient has at least one systemic medication
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM systemic_medications sm 
                    WHERE sm.patient_id = ps.patient_id
                )
            ''')
        elif systemic_meds_filter == '0':
            # Patient has no systemic medications
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM systemic_medications sm 
                    WHERE sm.patient_id = ps.patient_id
                )
            ''')
        elif systemic_meds_filter == 'ND':
            # No systemic medications
            where_clauses.append('''
                NOT EXISTS (
                    SELECT 1 FROM systemic_medications sm 
                    WHERE sm.patient_id = ps.patient_id
                )
            ''')
        elif systemic_meds_filter == 'not_0_not_nd':
            # Patient has at least one systemic medication
            where_clauses.append('''
                EXISTS (
                    SELECT 1 FROM systemic_medications sm 
                    WHERE sm.patient_id = ps.patient_id
                )
            ''')

    # Build final WHERE clause
    if where_clauses:
        where_clause = ' AND ' + ' AND '.join(where_clauses)
    else:
        where_clause = ''

    return where_clause, params


# =============== BACKUP FUNCTIONS ===============
def load_backup_config():
    """Load backup configuration from file"""
    if os.path.exists(BACKUP_CONFIG_FILE):
        with open(BACKUP_CONFIG_FILE, 'r') as f:
            config = json.load(f)
            # Override with environment variables if set
            config['backup_dir'] = os.getenv('BACKUP_DIR', config.get('backup_dir', DEFAULT_BACKUP_DIR))
            config['retention_days'] = int(os.getenv('BACKUP_RETENTION_DAYS',
                                           config.get('retention_days', DEFAULT_RETENTION_DAYS)))
            return config
    return {
        'backup_dir': DEFAULT_BACKUP_DIR,
        'schedule': 'disabled',
        'retention_days': DEFAULT_RETENTION_DAYS,
        'auto_backup_enabled': False
    }


def save_backup_config(config):
    """Save backup configuration to file"""
    with open(BACKUP_CONFIG_FILE, 'w') as f:
        json.dump(config, f, indent=2)


def format_file_size(size):
    """Format file size in human-readable format"""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size < 1024.0:
            return f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{size:.1f} TB"


def scheduled_backup():
    """Function to run scheduled backups"""
    config = load_backup_config()
    backup_dir = config['backup_dir']

    # Create backup directory if it doesn't exist
    os.makedirs(backup_dir, exist_ok=True)

    # Generate backup filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = os.path.join(backup_dir, f'raman_scheduled_{timestamp}.sql')

    try:
        # Create pg_dump command
        env = os.environ.copy()
        env['PGPASSWORD'] = DB_CONFIG['password']

        cmd = [
            'pg_dump',
            '-h', DB_CONFIG['host'],
            '-p', DB_CONFIG['port'],
            '-U', DB_CONFIG['user'],
            '-d', DB_CONFIG['dbname'],
            '-f', backup_file,
            '--no-password'
        ]

        subprocess.run(cmd, env=env, capture_output=True)

        # Clean old backups
        clean_old_backups(config)

        print(f"Scheduled backup completed: {backup_file}")

    except Exception as e:
        print(f"Scheduled backup failed: {e}")


def clean_old_backups(config):
    """Remove backups older than retention period"""
    backup_dir = config['backup_dir']
    retention_days = config.get('retention_days', 90)

    if not os.path.exists(backup_dir):
        return

    cutoff_time = datetime.now() - timedelta(days=retention_days)

    for file in os.listdir(backup_dir):
        if file.startswith('raman_') and (file.endswith('.sql') or file.endswith('.dump')):
            file_path = os.path.join(backup_dir, file)
            file_time = datetime.fromtimestamp(os.path.getmtime(file_path))

            if file_time < cutoff_time:
                try:
                    os.remove(file_path)
                    print(f"Deleted old backup: {file}")
                except Exception as e:
                    print(f"Failed to delete old backup {file}: {e}")


def start_backup_scheduler(config):
    """Start the backup scheduler"""
    global scheduler_thread, scheduler_running

    # Stop existing scheduler if running
    stop_backup_scheduler()

    schedule_type = config.get('schedule', 'disabled')

    if schedule_type == 'disabled':
        return

    # Clear existing jobs
    schedule.clear()

    # Set up schedule based on configuration
    if schedule_type == 'hourly':
        schedule.every().hour.do(scheduled_backup)
    elif schedule_type == 'daily':
        schedule.every().day.at("02:00").do(scheduled_backup)
    elif schedule_type == 'weekly':
        schedule.every().monday.at("02:00").do(scheduled_backup)
    elif schedule_type == 'monthly':
        # Run on the 1st of each month
        schedule.every().day.at("02:00").do(lambda: scheduled_backup() if datetime.now().day == 1 else None)

    # Start scheduler thread
    scheduler_running = True
    scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
    scheduler_thread.start()
    print(f"Backup scheduler started: {schedule_type}")


def stop_backup_scheduler():
    """Stop the backup scheduler"""
    global scheduler_running
    scheduler_running = False
    schedule.clear()
    print("Backup scheduler stopped")


def run_scheduler():
    """Run the scheduler in a separate thread"""
    while scheduler_running:
        schedule.run_pending()
        time.sleep(60)  # Check every minute


# =============== ICD-10 INITIALIZATION FROM EXCEL ===============
def init_icd10_from_excel():
    """Initialize ICD-10 codes from Excel files if they exist"""
    conn = get_db_connection()
    if not conn:
        print("Failed to connect to database for ICD-10 initialization")
        return

    try:
        cur = conn.cursor()

        # Check if ocular codes table is empty
        cur.execute("SELECT COUNT(*) FROM icd10_ocular_conditions")
        ocular_count = cur.fetchone()[0]

        # Import ocular codes if table is empty and Excel file exists
        if ocular_count == 0 and os.path.exists('ICD10_eye_codes.xlsx'):
            print("Importing ICD-10 Ocular codes from Excel...")
            eye_df = pd.read_excel('ICD10_eye_codes.xlsx')

            for _, row in eye_df.iterrows():
                code = str(row['ICD-10 Code']).strip()
                description = str(row['Description']).strip()

                # Determine category based on code prefix
                category = None
                if code.startswith('H0'):
                    category = 'Eyelid, lacrimal system and orbit'
                elif code.startswith('H1'):
                    category = 'Conjunctiva and cornea'
                elif code.startswith('H2'):
                    category = 'Lens'
                elif code.startswith('H3'):
                    category = 'Choroid and retina'
                elif code.startswith('H4'):
                    category = 'Glaucoma and optic nerve'
                elif code.startswith('H5'):
                    category = 'Ocular muscles and visual disorders'

                cur.execute("""
                    INSERT INTO icd10_ocular_conditions (code, description, category, active)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (code) DO NOTHING
                """, (code, description, category))

            conn.commit()
            print(f"Imported {len(eye_df)} ocular ICD-10 codes")
        elif ocular_count == 0:
            # Fall back to sample codes if Excel not available
            print("Excel file not found, using sample ocular codes...")
            ocular_codes = [
                ('H25.9', 'Senile cataract, unspecified', 'Cataract'),
                ('H26.9', 'Cataract, unspecified', 'Cataract'),
                ('H40.1', 'Primary open-angle glaucoma', 'Glaucoma'),
                ('H40.2', 'Primary angle-closure glaucoma', 'Glaucoma'),
            ]
            cur.executemany('''
                INSERT INTO icd10_ocular_conditions (code, description, category, active)
                VALUES (%s, %s, %s, TRUE)
            ''', ocular_codes)
            conn.commit()

        # Check if systemic codes table is empty
        cur.execute("SELECT COUNT(*) FROM icd10_systemic_conditions")
        systemic_count = cur.fetchone()[0]

        # Import systemic codes if table is empty and Excel file exists
        if systemic_count == 0 and os.path.exists('ICD10_non_eye_codes.xlsx'):
            print("Importing ICD-10 Systemic codes from Excel...")
            non_eye_df = pd.read_excel('ICD10_non_eye_codes.xlsx')

            for _, row in non_eye_df.iterrows():
                code = str(row['ICD-10 Code']).strip()
                description = str(row['Description']).strip()

                # Determine category based on code prefix
                category = None
                if code.startswith('A') or code.startswith('B'):
                    category = 'Infectious diseases'
                elif code.startswith('C') or code.startswith('D'):
                    category = 'Neoplasms'
                elif code.startswith('E'):
                    category = 'Endocrine and metabolic'
                elif code.startswith('F'):
                    category = 'Mental disorders'
                elif code.startswith('G'):
                    category = 'Nervous system'
                elif code.startswith('I'):
                    category = 'Circulatory system'
                elif code.startswith('J'):
                    category = 'Respiratory system'
                elif code.startswith('K'):
                    category = 'Digestive system'
                elif code.startswith('L'):
                    category = 'Skin diseases'
                elif code.startswith('M'):
                    category = 'Musculoskeletal'
                elif code.startswith('N'):
                    category = 'Genitourinary'
                elif code.startswith('O'):
                    category = 'Pregnancy'
                elif code.startswith('P'):
                    category = 'Perinatal'
                elif code.startswith('Q'):
                    category = 'Congenital'
                elif code.startswith('R'):
                    category = 'Symptoms'
                elif code.startswith('S') or code.startswith('T'):
                    category = 'Injury'
                elif code.startswith('V') or code.startswith('W') or code.startswith('X') or code.startswith('Y'):
                    category = 'External causes'
                elif code.startswith('Z'):
                    category = 'Health factors'

                cur.execute("""
                    INSERT INTO icd10_systemic_conditions (code, description, category, active)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (code) DO NOTHING
                """, (code, description, category))

            conn.commit()
            print(f"Imported {len(non_eye_df)} systemic ICD-10 codes")
        elif systemic_count == 0:
            # Fall back to sample codes if Excel not available
            print("Excel file not found, using sample systemic codes...")
            systemic_codes = [
                ('E11.9', 'Type 2 diabetes mellitus without complications', 'Endocrine'),
                ('I10', 'Essential (primary) hypertension', 'Cardiovascular'),
                ('E78.5', 'Hyperlipidemia, unspecified', 'Endocrine'),
            ]
            cur.executemany('''
                INSERT INTO icd10_systemic_conditions (code, description, category, active)
                VALUES (%s, %s, %s, TRUE)
            ''', systemic_codes)
            conn.commit()

        cur.close()
        conn.close()

    except Exception as e:
        print(f"Error initializing ICD-10 codes: {e}")
        if conn:
            conn.rollback()
            conn.close()


# Dynamic Generic Component Extraction for reporting purposes
def get_all_generic_components():
    """
    Dynamically extract all unique generic components from the medications table
    Returns a set of unique generic names
    """
    conn = get_db_connection()
    if not conn:
        return set()

    try:
        cur = conn.cursor()

        # Get all generic names from medications table
        cur.execute('''
            SELECT DISTINCT generic_name 
            FROM medications 
            WHERE generic_name IS NOT NULL AND generic_name != ''
        ''')

        all_generics = set()

        for row in cur.fetchall():
            generic_name = row[0].strip()

            # Split by semicolon for multi-component medications
            components = [c.strip().lower() for c in generic_name.split(';')]

            # Add each component to the set
            for component in components:
                if component:
                    all_generics.add(component)

        cur.close()
        conn.close()

        return all_generics

    except Exception as e:
        print(f"Error getting generic components: {e}")
        if conn:
            conn.close()
        return set()


def extract_generic_components_dynamic(medications_list, all_generic_components):
    """
    Extract individual generic drug components from a patient's medications
    Returns a dictionary with binary flags for each component
    """
    # Initialize all components as False
    generic_flags = {component: False for component in all_generic_components}

    # Check each medication
    for med in medications_list:
        if med and 'generic_name' in med and med['generic_name']:
            generic_name = str(med['generic_name']).lower()

            # Split by semicolon for multiple generics
            components = [c.strip() for c in generic_name.split(';')]

            # Check each component against our known generics
            for component in components:
                if component in generic_flags:
                    generic_flags[component] = True

    # Convert to column format with safe names
    return {f'takes_{make_safe_column_name(key)}': (1 if value else 0)
            for key, value in generic_flags.items()}


def make_safe_column_name(name):
    """Convert a string to a safe column name"""
    # Replace special characters with underscores
    import re
    safe_name = re.sub(r'[^a-zA-Z0-9]', '_', str(name))
    # Remove multiple underscores
    safe_name = re.sub(r'_+', '_', safe_name)
    # Remove leading/trailing underscores
    safe_name = safe_name.strip('_')
    return safe_name.lower()


def initialize_app():
    """Initialize the application - runs once when module is loaded"""
    print("\n" + "=" * 60)
    print("RAMAN MEDICAL RESEARCH DATABASE - Starting Up")
    print("=" * 60 + "\n")

    # Step 1: Check/Create database
    print("Step 1: Checking database existence...")
    if not create_database_if_not_exists():
        print("\n✗ Failed to create/access database. Continuing anyway...")
        # Don't exit - let the app start but log the error

    # Step 2: Initialize tables
    print("\nStep 2: Initializing database tables...")
    if not init_database():
        print("\n✗ Warning: Database initialization had issues")
        print("The application may not work correctly.\n")
    else:
        print("\n✓ Database structure is ready!")

    # Step 3: Ensure reference data is populated (even if tables already exist)
    print("\nStep 3: Checking and populating reference data...")
    try:
        populate_reference_data()
    except Exception as e:
        print(f"⚠ Warning: Error populating reference data: {e}")
        import traceback
        traceback.print_exc()

    # Step 4: Initialize backup scheduler
    print("\nStep 4: Checking backup scheduler configuration...")

    # Check if we're in the main worker or first worker
    # In Gunicorn, workers are numbered starting from 1
    worker_id = os.environ.get('GUNICORN_WORKER_ID', '0')

    # Alternative: Use file lock to ensure only one scheduler
    if worker_id == '0' or not os.path.exists('/tmp/backup_scheduler.lock'):
        try:
            # Create lock file
            if not os.path.exists('/tmp/backup_scheduler.lock'):
                with open('/tmp/backup_scheduler.lock', 'w') as f:
                    f.write(str(os.getpid()))

            config = load_backup_config()
            if config.get('auto_backup_enabled'):
                start_backup_scheduler(config)
                print(f"✓ Backup scheduler started in worker {worker_id}: {config.get('schedule')}")
                print(f"  Backup directory: {config.get('backup_dir')}")
                print(f"  Retention: {config.get('retention_days')} days")
            else:
                print("✓ Automatic backups are disabled")
        except Exception as e:
            print(f"⚠ Warning: Could not initialize backup scheduler: {e}")
    else:
        print(f"ℹ Skipping scheduler initialization in worker {worker_id} (already running in another worker)")

    print("\n" + "=" * 60)
    print("✓ Application initialization complete!")
    print("=" * 60 + "\n")


# Initialize when module loads
_initialized = False

if not _initialized:
    try:
        initialize_app()
        _initialized = True
    except Exception as e:
        print(f"Critical initialization error: {e}")
        import traceback
        traceback.print_exc()
        # Still set as initialized to prevent retry loops
        _initialized = True


# Routes

@app.route('/')
def index():
    """Redirect to dashboard if logged in, otherwise to login"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """User login"""
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        conn = get_db_connection()
        if not conn:
            flash('Database connection error', 'error')
            return render_template('login.html')

        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)
            cur.execute('SELECT * FROM users WHERE username = %s', (username,))
            user = cur.fetchone()

            if user and bcrypt.check_password_hash(user['password_hash'], password):
                session['user_id'] = user['user_id']
                session['username'] = user['username']
                session['role'] = user['role']

                # Update last login
                cur.execute('UPDATE users SET last_login = CURRENT_TIMESTAMP WHERE user_id = %s', (user['user_id'],))
                conn.commit()

                cur.close()
                conn.close()

                flash(f'Welcome back, {username}!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Invalid username or password', 'error')
                cur.close()
                conn.close()
        except Exception as e:
            flash(f'Login error: {str(e)}', 'error')
            if conn:
                conn.close()

    return render_template('login.html')


@app.route('/logout')
def logout():
    """User logout"""
    session.clear()
    flash('You have been logged out successfully', 'success')
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    """Main dashboard"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return render_template('dashboard.html', stats={})

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Get statistics
        cur.execute('SELECT COUNT(*) as total FROM patients_sensitive')
        total_patients = cur.fetchone()['total']

        cur.execute('SELECT COUNT(*) as total FROM users')
        total_users = cur.fetchone()['total']

        # Get next available patient ID (based on actual database content)
        next_patient_id = get_next_available_patient_id()

        stats = {
            'total_patients': total_patients,
            'total_users': total_users,
            'next_patient_id': next_patient_id
        }

        cur.close()
        conn.close()

        return render_template('dashboard.html', stats=stats)
    except Exception as e:
        flash(f'Error loading dashboard: {str(e)}', 'error')
        if conn:
            conn.close()
        return render_template('dashboard.html', stats={})


@app.route('/api/check-patient-id/<int:patient_id>')
@staff_or_admin_required
def api_check_patient_id(patient_id):
    """API endpoint to check if patient ID exists"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed', 'exists': True}), 500

    try:
        cur = conn.cursor()
        cur.execute("SELECT COUNT(*) FROM patients_sensitive WHERE patient_id = %s", (patient_id,))
        count = cur.fetchone()[0]
        exists = count > 0

        cur.close()
        conn.close()

        return jsonify({
            'exists': exists,
            'patient_id': patient_id,
            'available': not exists
        })
    except Exception as e:
        print(f"[API] Error checking patient ID {patient_id}: {e}")
        if conn:
            conn.close()
        return jsonify({'error': str(e), 'exists': True}), 500


@app.route('/api/next-patient-id')
@staff_or_admin_required
def api_next_patient_id():
    """API endpoint to get next available patient ID"""
    next_id = get_next_available_patient_id()
    if next_id:
        return jsonify({'patient_id': next_id, 'available': True})
    return jsonify({'error': 'Could not generate patient ID'}), 500


# New Patient Route

@app.route('/new-patient', methods=['GET', 'POST'])
@staff_or_admin_required
def new_patient():
    """Create new patient record"""
    if request.method == 'GET':
        # Get reference data for dropdowns
        conn = get_db_connection()
        if not conn:
            flash('Database connection error', 'error')
            return redirect(url_for('dashboard'))

        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)

            # Get ICD-10 ocular conditions
            cur.execute('SELECT code, description FROM icd10_ocular_conditions WHERE active = TRUE ORDER BY code')
            icd10_ocular = cur.fetchall()

            # Get ICD-10 systemic conditions
            cur.execute('SELECT code, description FROM icd10_systemic_conditions WHERE active = TRUE ORDER BY code')
            icd10_systemic = cur.fetchall()

            # Get medications
            cur.execute(
                'SELECT trade_name, generic_name, medication_type FROM medications WHERE active = TRUE ORDER BY trade_name')
            medications = cur.fetchall()

            # Get surgeries
            cur.execute('SELECT code, description FROM surgeries WHERE active = TRUE ORDER BY code')
            surgeries = cur.fetchall()

            # Get next patient ID based on actual database content
            next_id = get_next_available_patient_id()

            cur.close()
            conn.close()

            # Prepare stats with default values (in case template needs them)
            stats = {
                'total_patients': 0,
                'total_users': 0,
                'next_patient_id': next_id
            }

            return render_template('new_patient.html',
                                   icd10_ocular=icd10_ocular,
                                   icd10_systemic=icd10_systemic,
                                   medications=medications,
                                   surgeries=surgeries,
                                   next_patient_id=next_id,
                                   stats=stats)
        except Exception as e:
            flash(f'Error loading form: {str(e)}', 'error')
            if conn:
                conn.close()
            return redirect(url_for('dashboard'))

    # POST - save new patient
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('new_patient'))

    try:
        cur = conn.cursor()

        # Get form data - General Data
        patient_id = int(request.form.get('patient_id'))
        patient_name = request.form.get('patient_name')
        mbo = request.form.get('mbo')
        sex = request.form.get('sex')

        # Parse dates from separate day/month/year fields
        dob_day = request.form.get('dob_day')
        dob_month = request.form.get('dob_month')
        dob_year = request.form.get('dob_year')
        if dob_day and dob_month and dob_year:
            date_of_birth = date(int(dob_year), int(dob_month), int(dob_day))
        else:
            date_of_birth = None

        collection_day = request.form.get('collection_day')
        collection_month = request.form.get('collection_month')
        collection_year = request.form.get('collection_year')
        if collection_day and collection_month and collection_year:
            date_of_sample_collection = date(int(collection_year), int(collection_month), int(collection_day))
        else:
            date_of_sample_collection = None

        eye = request.form.get('eye')

        # Check if patient ID already exists
        if check_patient_id_exists(patient_id):
            flash(f'Patient ID {patient_id} already exists. Please use a different ID.', 'error')
            conn.close()
            return redirect(url_for('new_patient'))

        # Generate person hash and calculate age
        person_hash = generate_person_hash(mbo)
        age = calculate_age(date_of_birth, date_of_sample_collection)

        # Insert into patients_sensitive
        cur.execute('''
            INSERT INTO patients_sensitive (patient_id, patient_name, mbo, date_of_birth, date_of_sample_collection)
            VALUES (%s, %s, %s, %s, %s)
        ''', (patient_id, patient_name, mbo, date_of_birth, date_of_sample_collection))

        # Insert into patients_statistical
        cur.execute('''
            INSERT INTO patients_statistical (patient_id, person_hash, age, sex, eye)
            VALUES (%s, %s, %s, %s, %s)
        ''', (patient_id, person_hash, age, sex, eye))

        # Main Ocular Conditions
        lens_status = request.form.get('lens_status', 'ND')
        locs_iii_no = request.form.get('locs_no', 'ND')
        locs_iii_nc = request.form.get('locs_nc', 'ND')
        locs_iii_c = request.form.get('locs_c', 'ND')
        locs_iii_p = request.form.get('locs_p', 'ND')
        iol_type = request.form.get('iol_type', 'ND')
        etiology_aphakia = request.form.get('aphakia_etiology', 'ND')

        glaucoma = request.form.get('glaucoma', 'ND')
        oht_or_pac = request.form.get('oht_or_pac', 'ND')
        etiology_glaucoma = request.form.get('glaucoma_etiology', 'ND')
        steroid_responder = request.form.get('steroid_responder', 'ND')
        pxs = request.form.get('pxs', '0')
        pds = request.form.get('pds', '0')

        diabetic_retinopathy = request.form.get('diabetic_retinopathy', '0')
        stage_diabetic_retinopathy = request.form.get('dr_stage', 'ND')
        stage_npdr = request.form.get('npdr_stage', 'ND')
        stage_pdr = request.form.get('pdr_stage', 'ND')

        macular_edema = request.form.get('macular_edema', '0')
        etiology_macular_edema = request.form.get('me_etiology', 'ND')

        macular_degeneration_dystrophy = request.form.get('macular_degeneration', '0')
        etiology_macular_deg_dyst = request.form.get('md_etiology', 'ND')
        stage_amd = request.form.get('amd_stage', 'ND')
        exudation_amd = request.form.get('amd_exudation', 'ND')
        stage_other_macular_deg = request.form.get('other_md_stage', 'ND')
        exudation_other_macular_deg = request.form.get('other_md_exudation', 'ND')

        macular_hole_vmt = request.form.get('mh_vmt', '0')
        etiology_mh_vmt = request.form.get('mh_vmt_etiology', 'ND')
        cause_secondary_mh_vmt = request.form.get('secondary_mh_vmt_cause', 'ND')
        treatment_status_mh_vmt = request.form.get('mh_vmt_treatment_status', 'ND')

        epiretinal_membrane = request.form.get('epiretinal_membrane', '0')
        etiology_erm = request.form.get('erm_etiology', 'ND')
        cause_secondary_erm = request.form.get('secondary_erm_cause', 'ND')
        treatment_status_erm = request.form.get('erm_treatment_status', 'ND')

        retinal_detachment = request.form.get('retinal_detachment', '0')
        etiology_rd = request.form.get('rd_etiology', 'ND')
        treatment_status_rd = request.form.get('rd_treatment_status', 'ND')
        pvr = request.form.get('pvr', 'ND')

        vitreous_haemorrhage_opacification = request.form.get('vitreous_opacification', '0')
        etiology_vitreous_haemorrhage = request.form.get('vh_etiology', 'ND')

        # Insert ocular conditions
        cur.execute('''
            INSERT INTO ocular_conditions (
                patient_id, lens_status, locs_iii_no, locs_iii_nc, locs_iii_c, locs_iii_p,
                iol_type, etiology_aphakia, glaucoma, oht_or_pac, etiology_glaucoma,
                steroid_responder, pxs, pds, diabetic_retinopathy, stage_diabetic_retinopathy,
                stage_npdr, stage_pdr, macular_edema, etiology_macular_edema,
                macular_degeneration_dystrophy, etiology_macular_deg_dyst, stage_amd, exudation_amd,
                stage_other_macular_deg, exudation_other_macular_deg, macular_hole_vmt, etiology_mh_vmt,
                cause_secondary_mh_vmt, treatment_status_mh_vmt, epiretinal_membrane, etiology_erm,
                cause_secondary_erm, treatment_status_erm, retinal_detachment, etiology_rd,
                treatment_status_rd, pvr, vitreous_haemorrhage_opacification, etiology_vitreous_haemorrhage
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (patient_id, lens_status, locs_iii_no, locs_iii_nc, locs_iii_c, locs_iii_p,
              iol_type, etiology_aphakia, glaucoma, oht_or_pac, etiology_glaucoma,
              steroid_responder, pxs, pds, diabetic_retinopathy, stage_diabetic_retinopathy,
              stage_npdr, stage_pdr, macular_edema, etiology_macular_edema,
              macular_degeneration_dystrophy, etiology_macular_deg_dyst, stage_amd, exudation_amd,
              stage_other_macular_deg, exudation_other_macular_deg, macular_hole_vmt, etiology_mh_vmt,
              cause_secondary_mh_vmt, treatment_status_mh_vmt, epiretinal_membrane, etiology_erm,
              cause_secondary_erm, treatment_status_erm, retinal_detachment, etiology_rd,
              treatment_status_rd, pvr, vitreous_haemorrhage_opacification, etiology_vitreous_haemorrhage))

        # Other Ocular Conditions (multiple entries possible)
        other_ocular_conditions = request.form.getlist('other_ocular_condition[]')
        other_ocular_eyes = request.form.getlist('other_ocular_condition_eye[]')
        for icd10_code, eye_affected in zip(other_ocular_conditions, other_ocular_eyes):
            if icd10_code and icd10_code not in ['0', 'ND']:
                cur.execute('''
                    INSERT INTO other_ocular_conditions (patient_id, icd10_code, eye)
                    VALUES (%s, %s, %s)
                ''', (patient_id, icd10_code, eye_affected))

        # Previous Ocular Surgeries (multiple entries possible)
        surgeries_list = request.form.getlist('previous_surgery[]')
        surgeries_eyes = request.form.getlist('previous_surgery_eye[]')
        for surgery_code, eye_affected in zip(surgeries_list, surgeries_eyes):
            if surgery_code and surgery_code not in ['0', 'ND']:
                cur.execute('''
                    INSERT INTO previous_ocular_surgeries (patient_id, surgery_code, eye)
                    VALUES (%s, %s, %s)
                ''', (patient_id, surgery_code, eye_affected))

        # Systemic Conditions (multiple entries possible)
        systemic_conditions_list = request.form.getlist('systemic_condition[]')
        for icd10_code in systemic_conditions_list:
            if icd10_code and icd10_code not in ['0', 'ND']:
                cur.execute('''
                    INSERT INTO systemic_conditions (patient_id, icd10_code)
                    VALUES (%s, %s)
                ''', (patient_id, icd10_code))

        # Ocular Medications (multiple entries possible)
        ocular_meds_list = request.form.getlist('ocular_medication[]')
        ocular_meds_eyes = request.form.getlist('ocular_medication_eye[]')
        ocular_meds_days = request.form.getlist('ocular_medication_days[]')

        for medication, eye_affected, last_app in zip(ocular_meds_list, ocular_meds_eyes, ocular_meds_days):
            if medication and medication not in ['0', 'ND']:
                # Split medication into trade_name|generic_name
                parts = medication.split('|')
                if len(parts) == 2:
                    trade_name, generic_name = parts
                    # Default to 0 if blank or invalid
                    if last_app and last_app.strip() and last_app.isdigit():
                        last_application_days = int(last_app)
                    else:
                        last_application_days = 0  # Default to 0
                    cur.execute('''
                        INSERT INTO ocular_medications (patient_id, trade_name, generic_name, eye, last_application_days)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (patient_id, trade_name, generic_name, eye_affected, last_application_days))

        # Systemic Medications (multiple entries possible)
        systemic_meds_list = request.form.getlist('systemic_medication[]')
        systemic_meds_days = request.form.getlist('systemic_medication_days[]')

        for medication, last_app in zip(systemic_meds_list, systemic_meds_days):
            if medication and medication not in ['0', 'ND']:
                # Split medication into trade_name|generic_name
                parts = medication.split('|')
                if len(parts) == 2:
                    trade_name, generic_name = parts
                    # Default to 0 if blank or invalid
                    if last_app and last_app.strip() and last_app.isdigit():
                        last_application_days = int(last_app)
                    else:
                        last_application_days = 0  # Default to 0
                    cur.execute('''
                        INSERT INTO systemic_medications (patient_id, trade_name, generic_name, last_application_days)
                        VALUES (%s, %s, %s, %s)
                    ''', (patient_id, trade_name, generic_name, last_application_days))

        conn.commit()
        cur.close()
        conn.close()

        flash(f'Patient #{patient_id:05d} - {patient_name} has been added successfully!', 'success')
        return redirect(url_for('dashboard'))

    except Exception as e:
        conn.rollback()
        flash(f'Error saving patient: {str(e)}', 'error')
        if conn:
            conn.close()
        return redirect(url_for('new_patient'))


# Validate Data / Edit Patient Routes

@app.route('/validate-data', methods=['GET', 'POST'])
@staff_or_admin_required
def validate_data():
    """Search and list patients for validation with optional filtering"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return render_template('validate_data.html', patients=[])

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Get search parameters (from GET for search, POST for filters)
        search_type = request.args.get('type', 'id')
        search_query = request.args.get('q', '')

        # Check if filters are being used (POST request with filters)
        using_filters = request.method == 'POST' and request.form.get('use_filters') == '1'

        # Base query
        base_query = '''
            SELECT ps.patient_id, ps.patient_name, ps.date_of_birth, ps.date_of_sample_collection,
                   pst.sex, pst.eye
            FROM patients_sensitive ps
            JOIN patients_statistical pst ON ps.patient_id = pst.patient_id
        '''

        params = []
        where_clauses = []

        if using_filters:
            # Use the existing build_filter_clause function for filtering
            # Need to join with ocular_conditions table for filter support
            base_query = '''
                SELECT DISTINCT ps.patient_id, ps.patient_name, ps.date_of_birth, 
                       ps.date_of_sample_collection, pst.sex, pst.eye
                FROM patients_sensitive ps
                JOIN patients_statistical pst ON ps.patient_id = pst.patient_id
                LEFT JOIN ocular_conditions oc ON ps.patient_id = oc.patient_id
                WHERE 1=1
            '''

            # Build filter clause using existing function
            filter_clause, filter_params = build_filter_clause(request.form)
            base_query += filter_clause
            params.extend(filter_params)

            # Add search query on top of filters if provided
            if search_query:
                if search_type == 'id':
                    base_query += ' AND CAST(ps.patient_id AS TEXT) LIKE %s'
                    params.append(f'%{search_query}%')
                elif search_type == 'name':
                    base_query += ' AND LOWER(ps.patient_name) LIKE LOWER(%s)'
                    params.append(f'%{search_query}%')
                elif search_type == 'mbo':
                    base_query += ' AND ps.mbo LIKE %s'
                    params.append(f'%{search_query}%')

            base_query += ' ORDER BY ps.patient_id DESC LIMIT 100'

        elif search_query:
            # Traditional search without filters
            if search_type == 'id':
                base_query += '''
                    WHERE CAST(ps.patient_id AS TEXT) LIKE %s
                    ORDER BY ps.patient_id DESC
                    LIMIT 20
                '''
                params.append(f'%{search_query}%')
            elif search_type == 'name':
                base_query += '''
                    WHERE LOWER(ps.patient_name) LIKE LOWER(%s)
                    ORDER BY ps.patient_id DESC
                    LIMIT 20
                '''
                params.append(f'%{search_query}%')
            elif search_type == 'mbo':
                base_query += '''
                    WHERE ps.mbo LIKE %s
                    ORDER BY ps.patient_id DESC
                    LIMIT 20
                '''
                params.append(f'%{search_query}%')
        else:
            # Show 20 most recent patients if no search query or filters
            base_query += 'ORDER BY ps.patient_id DESC LIMIT 20'

        # Execute query
        if params:
            cur.execute(base_query, params)
        else:
            cur.execute(base_query)

        patients = cur.fetchall()

        cur.close()
        conn.close()

        return render_template('validate_data.html',
                               patients=patients,
                               search_type=search_type,
                               search_query=search_query,
                               using_filters=using_filters,
                               filters=request.form if using_filters else {})

    except Exception as e:
        flash(f'Error searching patients: {str(e)}', 'error')
        if conn:
            conn.close()
        return render_template('validate_data.html',
                               patients=[],
                               search_type=search_type,
                               search_query=search_query)


@app.route('/edit-patient/<int:patient_id>', methods=['GET', 'POST'])
@staff_or_admin_required
def edit_patient(patient_id):
    """Edit existing patient record"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('validate_data'))

    if request.method == 'GET':
        # Load patient data and reference lists
        try:
            cur = conn.cursor(cursor_factory=RealDictCursor)

            # Get patient data
            cur.execute('''
                SELECT ps.*, pst.sex, pst.eye, pst.age
                FROM patients_sensitive ps
                JOIN patients_statistical pst ON ps.patient_id = pst.patient_id
                WHERE ps.patient_id = %s
            ''', (patient_id,))
            patient = cur.fetchone()

            if not patient:
                flash(f'Patient #{patient_id} not found', 'error')
                cur.close()
                conn.close()
                return redirect(url_for('validate_data'))

            # Get ocular conditions
            cur.execute('SELECT * FROM ocular_conditions WHERE patient_id = %s', (patient_id,))
            ocular_conditions = cur.fetchone()

            # Get other ocular conditions
            cur.execute('SELECT * FROM other_ocular_conditions WHERE patient_id = %s', (patient_id,))
            other_ocular = cur.fetchall()

            # Get previous surgeries
            cur.execute('SELECT * FROM previous_ocular_surgeries WHERE patient_id = %s', (patient_id,))
            surgeries = cur.fetchall()

            # Get systemic conditions
            cur.execute('SELECT * FROM systemic_conditions WHERE patient_id = %s', (patient_id,))
            systemic = cur.fetchall()

            # Get ocular medications
            cur.execute('SELECT * FROM ocular_medications WHERE patient_id = %s', (patient_id,))
            ocular_meds = cur.fetchall()

            # Get systemic medications
            cur.execute('SELECT * FROM systemic_medications WHERE patient_id = %s', (patient_id,))
            systemic_meds = cur.fetchall()

            # Get reference data for dropdowns
            cur.execute('SELECT code, description FROM icd10_ocular_conditions WHERE active = TRUE ORDER BY code')
            icd10_ocular = cur.fetchall()

            cur.execute('SELECT code, description FROM icd10_systemic_conditions WHERE active = TRUE ORDER BY code')
            icd10_systemic = cur.fetchall()

            cur.execute(
                'SELECT trade_name, generic_name, medication_type FROM medications WHERE active = TRUE ORDER BY trade_name')
            medications = cur.fetchall()

            cur.execute('SELECT code, description FROM surgeries WHERE active = TRUE ORDER BY code')
            surgeries_list = cur.fetchall()

            cur.close()
            conn.close()

            # Prepare stats with default values (in case template needs them)
            stats = {
                'total_patients': 0,
                'total_users': 0,
                'next_patient_id': STARTING_PATIENT_ID
            }

            return render_template('edit_patient.html',
                                   patient=patient,
                                   ocular_conditions=ocular_conditions,
                                   other_conditions=other_ocular,
                                   surgeries=surgeries,
                                   systemic=systemic,
                                   ocular_meds=ocular_meds,
                                   systemic_meds=systemic_meds,
                                   icd10_ocular=icd10_ocular,
                                   icd10_systemic=icd10_systemic,
                                   medications=medications,
                                   surgeries_list=surgeries_list,
                                   stats=stats)
        except Exception as e:
            flash(f'Error loading patient data: {str(e)}', 'error')
            if conn:
                conn.close()
            return redirect(url_for('validate_data'))

    # POST - Update patient data
    try:
        cur = conn.cursor()

        # Get basic patient data
        patient_name = request.form.get('patient_name')
        mbo = request.form.get('mbo')
        sex = request.form.get('sex')
        eye = request.form.get('eye')

        # Get date fields
        dob_day = request.form.get('dob_day')
        dob_month = request.form.get('dob_month')
        dob_year = request.form.get('dob_year')
        dosc_day = request.form.get('collection_day')
        dosc_month = request.form.get('collection_month')
        dosc_year = request.form.get('collection_year')

        # Parse dates
        date_of_birth = None
        if dob_day and dob_month and dob_year:
            try:
                date_of_birth = date(int(dob_year), int(dob_month), int(dob_day))
            except ValueError:
                flash('Invalid date of birth', 'error')
                return redirect(url_for('edit_patient', patient_id=patient_id))

        date_of_sample_collection = None
        if dosc_day and dosc_month and dosc_year:
            try:
                date_of_sample_collection = date(int(dosc_year), int(dosc_month), int(dosc_day))
            except ValueError:
                flash('Invalid sample collection date', 'error')
                return redirect(url_for('edit_patient', patient_id=patient_id))

        # Calculate age and person hash
        age = None
        if date_of_birth and date_of_sample_collection:
            age = date_of_sample_collection.year - date_of_birth.year
            if date_of_sample_collection.month < date_of_birth.month or \
                    (date_of_sample_collection.month == date_of_birth.month and
                     date_of_sample_collection.day < date_of_birth.day):
                age -= 1

        person_hash = hashlib.sha256(mbo.encode()).hexdigest() if mbo else None

        # Update patients_sensitive table
        cur.execute('''
            UPDATE patients_sensitive
            SET patient_name = %s, mbo = %s, date_of_birth = %s, 
                date_of_sample_collection = %s, updated_at = CURRENT_TIMESTAMP
            WHERE patient_id = %s
        ''', (patient_name, mbo, date_of_birth, date_of_sample_collection, patient_id))

        # Update patients_statistical table
        cur.execute('''
            UPDATE patients_statistical
            SET person_hash = %s, age = %s, sex = %s, eye = %s
            WHERE patient_id = %s
        ''', (person_hash, age, sex, eye, patient_id))

        # Get all ocular condition fields
        lens_status = request.form.get('lens_status', 'ND')
        locs_iii_no = request.form.get('locs_no', 'ND')
        locs_iii_nc = request.form.get('locs_nc', 'ND')
        locs_iii_c = request.form.get('locs_c', 'ND')
        locs_iii_p = request.form.get('locs_p', 'ND')
        iol_type = request.form.get('iol_type', 'ND')
        etiology_aphakia = request.form.get('aphakia_etiology', 'ND')
        glaucoma = request.form.get('glaucoma', '0')
        oht_or_pac = request.form.get('oht_or_pac', 'ND')
        etiology_glaucoma = request.form.get('glaucoma_etiology', 'ND')
        steroid_responder = request.form.get('steroid_responder', 'ND')
        pxs = request.form.get('pxs', '0')
        pds = request.form.get('pds', '0')
        diabetic_retinopathy = request.form.get('diabetic_retinopathy', '0')
        stage_diabetic_retinopathy = request.form.get('dr_stage', 'ND')
        stage_npdr = request.form.get('npdr_stage', 'ND')
        stage_pdr = request.form.get('pdr_stage', 'ND')
        macular_edema = request.form.get('macular_edema', '0')
        etiology_macular_edema = request.form.get('me_etiology', 'ND')
        macular_degeneration_dystrophy = request.form.get('macular_degeneration', '0')
        etiology_macular_deg_dyst = request.form.get('md_etiology', 'ND')
        stage_amd = request.form.get('amd_stage', 'ND')
        exudation_amd = request.form.get('amd_exudation', 'ND')
        stage_other_macular_deg = request.form.get('other_md_stage', 'ND')
        exudation_other_macular_deg = request.form.get('other_md_exudation', 'ND')
        macular_hole_vmt = request.form.get('mh_vmt', '0')
        etiology_mh_vmt = request.form.get('mh_vmt_etiology', 'ND')
        cause_secondary_mh_vmt = request.form.get('secondary_mh_vmt_cause', 'ND')
        treatment_status_mh_vmt = request.form.get('mh_vmt_treatment_status', 'ND')
        epiretinal_membrane = request.form.get('epiretinal_membrane', '0')
        etiology_erm = request.form.get('erm_etiology', 'ND')
        cause_secondary_erm = request.form.get('secondary_erm_cause', 'ND')
        treatment_status_erm = request.form.get('erm_treatment_status', 'ND')
        retinal_detachment = request.form.get('retinal_detachment', '0')
        etiology_rd = request.form.get('rd_etiology', 'ND')
        treatment_status_rd = request.form.get('rd_treatment_status', 'ND')
        pvr = request.form.get('pvr', 'ND')
        vitreous_haemorrhage_opacification = request.form.get('vitreous_opacification', '0')
        etiology_vitreous_haemorrhage = request.form.get('vh_etiology', 'ND')

        # Update ocular_conditions table
        cur.execute('''
            UPDATE ocular_conditions
            SET lens_status = %s, locs_iii_no = %s, locs_iii_nc = %s, locs_iii_c = %s, locs_iii_p = %s,
                iol_type = %s, etiology_aphakia = %s, glaucoma = %s, oht_or_pac = %s, etiology_glaucoma = %s,
                steroid_responder = %s, pxs = %s, pds = %s, diabetic_retinopathy = %s,
                stage_diabetic_retinopathy = %s, stage_npdr = %s, stage_pdr = %s, macular_edema = %s,
                etiology_macular_edema = %s, macular_degeneration_dystrophy = %s,
                etiology_macular_deg_dyst = %s, stage_amd = %s, exudation_amd = %s,
                stage_other_macular_deg = %s, exudation_other_macular_deg = %s, macular_hole_vmt = %s,
                etiology_mh_vmt = %s, cause_secondary_mh_vmt = %s, treatment_status_mh_vmt = %s,
                epiretinal_membrane = %s, etiology_erm = %s, cause_secondary_erm = %s,
                treatment_status_erm = %s, retinal_detachment = %s, etiology_rd = %s,
                treatment_status_rd = %s, pvr = %s, vitreous_haemorrhage_opacification = %s,
                etiology_vitreous_haemorrhage = %s, updated_at = CURRENT_TIMESTAMP
            WHERE patient_id = %s
        ''', (lens_status, locs_iii_no, locs_iii_nc, locs_iii_c, locs_iii_p,
              iol_type, etiology_aphakia, glaucoma, oht_or_pac, etiology_glaucoma,
              steroid_responder, pxs, pds, diabetic_retinopathy, stage_diabetic_retinopathy,
              stage_npdr, stage_pdr, macular_edema, etiology_macular_edema,
              macular_degeneration_dystrophy, etiology_macular_deg_dyst, stage_amd, exudation_amd,
              stage_other_macular_deg, exudation_other_macular_deg, macular_hole_vmt, etiology_mh_vmt,
              cause_secondary_mh_vmt, treatment_status_mh_vmt, epiretinal_membrane, etiology_erm,
              cause_secondary_erm, treatment_status_erm, retinal_detachment, etiology_rd,
              treatment_status_rd, pvr, vitreous_haemorrhage_opacification, etiology_vitreous_haemorrhage,
              patient_id))

        # Delete existing many-to-many relationships and re-insert
        cur.execute('DELETE FROM other_ocular_conditions WHERE patient_id = %s', (patient_id,))
        cur.execute('DELETE FROM previous_ocular_surgeries WHERE patient_id = %s', (patient_id,))
        cur.execute('DELETE FROM systemic_conditions WHERE patient_id = %s', (patient_id,))
        cur.execute('DELETE FROM ocular_medications WHERE patient_id = %s', (patient_id,))
        cur.execute('DELETE FROM systemic_medications WHERE patient_id = %s', (patient_id,))

        # Other Ocular Conditions (multiple entries possible)
        other_ocular_conditions = request.form.getlist('other_ocular_condition[]')
        other_ocular_eyes = request.form.getlist('other_ocular_condition_eye[]')
        for icd10_code, eye_affected in zip(other_ocular_conditions, other_ocular_eyes):
            if icd10_code and icd10_code not in ['0', 'ND']:
                cur.execute('''
                    INSERT INTO other_ocular_conditions (patient_id, icd10_code, eye)
                    VALUES (%s, %s, %s)
                ''', (patient_id, icd10_code, eye_affected))

        # Previous Ocular Surgeries (multiple entries possible)
        surgeries_list = request.form.getlist('previous_surgery[]')
        surgeries_eyes = request.form.getlist('previous_surgery_eye[]')
        for surgery_code, eye_affected in zip(surgeries_list, surgeries_eyes):
            if surgery_code and surgery_code not in ['0', 'ND']:
                cur.execute('''
                    INSERT INTO previous_ocular_surgeries (patient_id, surgery_code, eye)
                    VALUES (%s, %s, %s)
                ''', (patient_id, surgery_code, eye_affected))

        # Systemic Conditions (multiple entries possible)
        systemic_conditions_list = request.form.getlist('systemic_condition[]')
        for icd10_code in systemic_conditions_list:
            if icd10_code and icd10_code not in ['0', 'ND']:
                cur.execute('''
                    INSERT INTO systemic_conditions (patient_id, icd10_code)
                    VALUES (%s, %s)
                ''', (patient_id, icd10_code))

        # Ocular Medications (multiple entries possible)
        ocular_meds_list = request.form.getlist('ocular_medication[]')
        ocular_meds_eyes = request.form.getlist('ocular_medication_eye[]')
        ocular_meds_days = request.form.getlist('ocular_medication_days[]')

        for medication, eye_affected, last_app in zip(ocular_meds_list, ocular_meds_eyes, ocular_meds_days):
            if medication and medication not in ['0', 'ND']:
                # Split medication into trade_name|generic_name
                parts = medication.split('|')
                if len(parts) == 2:
                    trade_name, generic_name = parts
                    # Default to 0 if blank or invalid
                    if last_app and last_app.strip() and last_app.isdigit():
                        last_application_days = int(last_app)
                    else:
                        last_application_days = 0  # Default to 0
                    cur.execute('''
                        INSERT INTO ocular_medications (patient_id, trade_name, generic_name, eye, last_application_days)
                        VALUES (%s, %s, %s, %s, %s)
                    ''', (patient_id, trade_name, generic_name, eye_affected, last_application_days))

        # Systemic Medications (multiple entries possible)
        systemic_meds_list = request.form.getlist('systemic_medication[]')
        systemic_meds_days = request.form.getlist('systemic_medication_days[]')

        for medication, last_app in zip(systemic_meds_list, systemic_meds_days):
            if medication and medication not in ['0', 'ND']:
                # Split medication into trade_name|generic_name
                parts = medication.split('|')
                if len(parts) == 2:
                    trade_name, generic_name = parts
                    # Default to 0 if blank or invalid
                    if last_app and last_app.strip() and last_app.isdigit():
                        last_application_days = int(last_app)
                    else:
                        last_application_days = 0  # Default to 0
                    cur.execute('''
                        INSERT INTO systemic_medications (patient_id, trade_name, generic_name, last_application_days)
                        VALUES (%s, %s, %s, %s)
                    ''', (patient_id, trade_name, generic_name, last_application_days))

        conn.commit()
        cur.close()
        conn.close()

        flash(f'Patient #{patient_id:05d} - {patient_name} has been updated successfully!', 'success')
        return redirect(url_for('validate_data'))

    except Exception as e:
        conn.rollback()
        flash(f'Error updating patient: {str(e)}', 'error')
        if conn:
            conn.close()
        return redirect(url_for('edit_patient', patient_id=patient_id))


# Export Data Route

@app.route('/export_data', methods=['GET', 'POST'])
@login_required
def export_data():
    """Export statistical data with BINARY COLUMNS - each medication/condition/surgery gets its own column"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return render_template('export_data.html',
                               stats={'total_patients': 0, 'gender': {'M': 0, 'F': 0}, 'age_distribution': []})

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Get statistics for display
        cur.execute('SELECT COUNT(*) as total FROM patients_sensitive')
        total_patients = cur.fetchone()['total']

        cur.execute('''
            SELECT sex, COUNT(*) as count
            FROM patients_statistical
            WHERE sex IN ('M', 'F')
            GROUP BY sex
        ''')
        gender_stats = cur.fetchall()
        gender = {'M': 0, 'F': 0}
        for row in gender_stats:
            gender[row['sex']] = row['count']

        # Age distribution
        cur.execute('''
            SELECT 
                CASE 
                    WHEN age < 18 THEN '0-17'
                    WHEN age < 30 THEN '18-29'
                    WHEN age < 40 THEN '30-39'
                    WHEN age < 50 THEN '40-49'
                    WHEN age < 60 THEN '50-59'
                    WHEN age < 70 THEN '60-69'
                    WHEN age < 80 THEN '70-79'
                    ELSE '80+'
                END as age_group,
                COUNT(*) as count
            FROM patients_statistical
            WHERE age IS NOT NULL
            GROUP BY age_group
            ORDER BY age_group
        ''')
        age_distribution = [(row['age_group'], row['count']) for row in cur.fetchall()]

        stats = {
            'total_patients': total_patients,
            'gender': gender,
            'age_distribution': age_distribution
        }

        if request.method == 'POST':
            # Handle export
            export_format = request.form.get('format', 'csv')
            data_type = request.form.get('data_type', 'anonymized')

            # Staff can only export anonymized data
            if session.get('role') == 'Staff':
                data_type = 'anonymized'

            # Get date range if provided
            date_from = request.form.get('date_from')
            date_to = request.form.get('date_to')

            # Get data inclusion options
            include_conditions = 'include_conditions' in request.form
            include_other_conditions = 'include_other_conditions' in request.form
            include_surgeries = 'include_surgeries' in request.form
            include_systemic = 'include_systemic' in request.form
            include_medications = 'include_medications' in request.form

            # ============================================================
            # STEP 1: Query all reference data (including inactive items used by any patient)
            # ============================================================

            # Get all medications (active OR used by patients)
            cur.execute('''
                SELECT DISTINCT generic_name
                FROM (
                    SELECT generic_name FROM medications WHERE active = TRUE
                    UNION
                    SELECT DISTINCT generic_name FROM ocular_medications
                    UNION
                    SELECT DISTINCT generic_name FROM systemic_medications
                ) AS all_meds
                ORDER BY generic_name
            ''')
            all_medications = [row['generic_name'] for row in cur.fetchall()]

            # Get all unique generic components for dynamic columns
            all_generic_components = get_all_generic_components()

            # Sort them alphabetically for consistent column ordering
            sorted_generic_components = sorted(all_generic_components)

            # Get all ocular ICD-10 codes (active OR used by patients)
            cur.execute('''
                SELECT DISTINCT code
                FROM (
                    SELECT code FROM icd10_ocular_conditions WHERE active = TRUE
                    UNION
                    SELECT DISTINCT icd10_code AS code FROM other_ocular_conditions
                ) AS all_codes
                ORDER BY code
            ''')
            all_ocular_codes = [row['code'] for row in cur.fetchall()]

            # Get all systemic ICD-10 codes (active OR used by patients)
            cur.execute('''
                SELECT DISTINCT code
                FROM (
                    SELECT code FROM icd10_systemic_conditions WHERE active = TRUE
                    UNION
                    SELECT DISTINCT icd10_code AS code FROM systemic_conditions
                ) AS all_codes
                ORDER BY code
            ''')
            all_systemic_codes = [row['code'] for row in cur.fetchall()]

            # Get all surgery codes (active OR used by patients)
            # FIXED: surgeries table uses 'code' not 'surgery_code'
            cur.execute('''
                SELECT DISTINCT code
                FROM (
                    SELECT code FROM surgeries WHERE active = TRUE
                    UNION
                    SELECT DISTINCT surgery_code AS code FROM previous_ocular_surgeries
                ) AS all_surgeries
                ORDER BY code
            ''')
            all_surgeries = [row['code'] for row in cur.fetchall()]

            # ============================================================
            # STEP 2: Build base query for patients
            # ============================================================

            if data_type == 'sensitive' and session.get('role') == 'Administrator':
                # Sensitive export - includes names and MBO
                base_query = '''
                    SELECT 
                        ps.patient_id,
                        ps.patient_name,
                        ps.mbo,
                        pst.sex,
                        ps.date_of_birth,
                        ps.date_of_sample_collection,
                        pst.eye,
                        pst.person_hash,
                        pst.age
                '''
            else:
                # Anonymized export
                base_query = '''
                    SELECT 
                        ps.patient_id,
                        pst.person_hash,
                        pst.sex,
                        pst.eye,
                        pst.age
                '''

            # Add ocular conditions if requested
            if include_conditions:
                base_query += ''',
                    oc.lens_status,
                    oc.locs_iii_no,
                    oc.locs_iii_nc,
                    oc.locs_iii_c,
                    oc.locs_iii_p,
                    oc.iol_type,
                    oc.etiology_aphakia,
                    oc.glaucoma,
                    oc.oht_or_pac,
                    oc.etiology_glaucoma,
                    oc.steroid_responder,
                    oc.pxs,
                    oc.pds,
                    oc.diabetic_retinopathy,
                    oc.stage_diabetic_retinopathy,
                    oc.stage_npdr,
                    oc.stage_pdr,
                    oc.macular_edema,
                    oc.etiology_macular_edema,
                    oc.macular_degeneration_dystrophy,
                    oc.etiology_macular_deg_dyst,
                    oc.stage_amd,
                    oc.exudation_amd,
                    oc.stage_other_macular_deg,
                    oc.exudation_other_macular_deg,
                    oc.macular_hole_vmt,
                    oc.etiology_mh_vmt,
                    oc.cause_secondary_mh_vmt,
                    oc.treatment_status_mh_vmt,
                    oc.epiretinal_membrane,
                    oc.etiology_erm,
                    oc.cause_secondary_erm,
                    oc.treatment_status_erm,
                    oc.retinal_detachment,
                    oc.etiology_rd,
                    oc.treatment_status_rd,
                    oc.pvr,
                    oc.vitreous_haemorrhage_opacification,
                    oc.etiology_vitreous_haemorrhage
                '''

            base_query += '''
                FROM patients_sensitive ps
                JOIN patients_statistical pst ON ps.patient_id = pst.patient_id
            '''

            if include_conditions:
                base_query += ' LEFT JOIN ocular_conditions oc ON ps.patient_id = oc.patient_id'

            base_query += ' WHERE 1=1'

            params = []

            # Add date filters
            if date_from:
                base_query += ' AND ps.date_of_sample_collection >= %s'
                params.append(date_from)
            if date_to:
                base_query += ' AND ps.date_of_sample_collection <= %s'
                params.append(date_to)

            # Add patient filters
            filter_clause, filter_params = build_filter_clause(request.form)
            base_query += filter_clause
            params.extend(filter_params)

            base_query += ' ORDER BY ps.patient_id'

            cur.execute(base_query, params)
            patients_data = cur.fetchall()

            # ============================================================
            # STEP 3: Preload all patient-related data for performance
            # ============================================================

            patient_ids = [p['patient_id'] for p in patients_data]

            # Preload other ocular conditions
            patient_ocular_conditions = {}
            if include_other_conditions and patient_ids:
                cur.execute('''
                    SELECT patient_id, icd10_code, eye 
                    FROM other_ocular_conditions 
                    WHERE patient_id = ANY(%s)
                ''', (patient_ids,))
                for row in cur.fetchall():
                    if row['patient_id'] not in patient_ocular_conditions:
                        patient_ocular_conditions[row['patient_id']] = []
                    patient_ocular_conditions[row['patient_id']].append(row)

            # Preload surgeries
            patient_surgeries = {}
            if include_surgeries and patient_ids:
                cur.execute('''
                    SELECT patient_id, surgery_code, eye 
                    FROM previous_ocular_surgeries 
                    WHERE patient_id = ANY(%s)
                ''', (patient_ids,))
                for row in cur.fetchall():
                    if row['patient_id'] not in patient_surgeries:
                        patient_surgeries[row['patient_id']] = []
                    patient_surgeries[row['patient_id']].append(row)

            # Preload systemic conditions
            patient_systemic = {}
            if include_systemic and patient_ids:
                cur.execute('''
                    SELECT patient_id, icd10_code 
                    FROM systemic_conditions 
                    WHERE patient_id = ANY(%s)
                ''', (patient_ids,))
                for row in cur.fetchall():
                    if row['patient_id'] not in patient_systemic:
                        patient_systemic[row['patient_id']] = []
                    patient_systemic[row['patient_id']].append(row)

            # Preload ocular medications
            patient_ocular_meds = {}
            if include_medications and patient_ids:
                cur.execute('''
                    SELECT patient_id, generic_name, eye, last_application_days 
                    FROM ocular_medications 
                    WHERE patient_id = ANY(%s)
                ''', (patient_ids,))
                for row in cur.fetchall():
                    if row['patient_id'] not in patient_ocular_meds:
                        patient_ocular_meds[row['patient_id']] = []
                    patient_ocular_meds[row['patient_id']].append(row)

            # Preload systemic medications
            patient_systemic_meds = {}
            if include_medications and patient_ids:
                cur.execute('''
                    SELECT patient_id, generic_name, last_application_days 
                    FROM systemic_medications 
                    WHERE patient_id = ANY(%s)
                ''', (patient_ids,))
                for row in cur.fetchall():
                    if row['patient_id'] not in patient_systemic_meds:
                        patient_systemic_meds[row['patient_id']] = []
                    patient_systemic_meds[row['patient_id']].append(row)

            # ============================================================
            # STEP 4: Build column headers (BINARY FORMAT)
            # ============================================================

            # Helper function to make safe column names
            def make_safe_column_name(name):
                """Convert any name to safe column name"""
                safe = str(name).lower()
                safe = safe.replace(' ', '_')
                safe = safe.replace('/', '_')
                safe = safe.replace('-', '_')
                safe = safe.replace('.', '_')
                safe = safe.replace('(', '')
                safe = safe.replace(')', '')
                safe = safe.replace('+', '_')
                safe = ''.join(c if c.isalnum() or c == '_' else '' for c in safe)
                # Ensure doesn't start with number
                if safe and safe[0].isdigit():
                    safe = 'x_' + safe
                return safe

            # Start with base columns
            if data_type == 'sensitive' and session.get('role') == 'Administrator':
                final_columns = [
                    'patient_id', 'patient_name', 'mbo', 'sex', 'date_of_birth',
                    'date_of_sample_collection', 'eye', 'person_hash', 'age'
                ]
            else:
                final_columns = ['patient_id', 'person_hash', 'sex', 'eye', 'age']

            # Add main condition columns if included
            if include_conditions:
                final_columns.extend([
                    'lens_status', 'locs_iii_no', 'locs_iii_nc', 'locs_iii_c', 'locs_iii_p',
                    'iol_type', 'etiology_aphakia', 'glaucoma', 'oht_or_pac', 'etiology_glaucoma',
                    'steroid_responder', 'pxs', 'pds', 'diabetic_retinopathy', 'stage_diabetic_retinopathy',
                    'stage_npdr', 'stage_pdr', 'macular_edema', 'etiology_macular_edema',
                    'macular_degeneration_dystrophy', 'etiology_macular_deg_dyst', 'stage_amd',
                    'exudation_amd', 'stage_other_macular_deg', 'exudation_other_macular_deg',
                    'macular_hole_vmt', 'etiology_mh_vmt', 'cause_secondary_mh_vmt',
                    'treatment_status_mh_vmt', 'epiretinal_membrane', 'etiology_erm',
                    'cause_secondary_erm', 'treatment_status_erm', 'retinal_detachment',
                    'etiology_rd', 'treatment_status_rd', 'pvr', 'vitreous_haemorrhage_opacification',
                    'etiology_vitreous_haemorrhage'
                ])

            # Add binary columns for other ocular conditions
            if include_other_conditions:
                for code in all_ocular_codes:
                    safe_code = make_safe_column_name(code)
                    final_columns.append(f'other_ocular_{safe_code}')
                    final_columns.append(f'other_ocular_{safe_code}_eye')

            # Add binary columns for surgeries
            if include_surgeries:
                for surgery in all_surgeries:
                    safe_surgery = make_safe_column_name(surgery)
                    final_columns.append(f'surgery_{safe_surgery}')
                    final_columns.append(f'surgery_{safe_surgery}_eye')

            # Add binary columns for systemic conditions
            if include_systemic:
                for code in all_systemic_codes:
                    safe_code = make_safe_column_name(code)
                    final_columns.append(f'systemic_{safe_code}')

            # Add binary columns for ocular medications
            if include_medications:
                for med in all_medications:
                    safe_med = make_safe_column_name(med)
                    final_columns.append(f'ocular_med_{safe_med}')
                    final_columns.append(f'ocular_med_{safe_med}_eye')
                    final_columns.append(f'ocular_med_{safe_med}_days')

            # Add binary columns for systemic medications
            if include_medications:
                for med in all_medications:
                    safe_med = make_safe_column_name(med)
                    final_columns.append(f'systemic_med_{safe_med}')
                    final_columns.append(f'systemic_med_{safe_med}_days')

            # Add binary columns for generic components
            if include_medications:
                for generic_component in sorted_generic_components:
                    safe_generic = make_safe_column_name(generic_component)
                    final_columns.append(f'takes_{safe_generic}')

            # ============================================================
            # STEP 5: Build export data with binary values
            # ============================================================

            export_data = []

            for patient in patients_data:
                # Initialize row with all columns set to default values
                row = {}

                # Fill base patient data
                for col in final_columns:
                    if col in patient:
                        value = patient[col]
                        # Convert dates to strings
                        if isinstance(value, (date, datetime)):
                            row[col] = value.strftime('%Y-%m-%d')
                        else:
                            row[col] = value
                    elif col.endswith('_eye'):
                        row[col] = 'ND'
                    elif col.endswith('_days'):
                        row[col] = 'ND'
                    elif col.startswith('other_ocular_') or col.startswith('surgery_') or \
                            col.startswith('systemic_') or col.startswith('ocular_med_') or \
                            col.startswith('systemic_med_'):
                        # Binary columns default to 0
                        if not col.endswith('_eye') and not col.endswith('_days'):
                            row[col] = 0
                        else:
                            row[col] = 'ND'
                    else:
                        row[col] = ''

                # Fill other ocular conditions (BINARY)
                if include_other_conditions:
                    for cond in patient_ocular_conditions.get(patient['patient_id'], []):
                        safe_code = make_safe_column_name(cond['icd10_code'])
                        row[f'other_ocular_{safe_code}'] = 1
                        row[f'other_ocular_{safe_code}_eye'] = cond['eye']

                # Fill surgeries (BINARY)
                if include_surgeries:
                    for surgery in patient_surgeries.get(patient['patient_id'], []):
                        safe_surgery = make_safe_column_name(surgery['surgery_code'])
                        row[f'surgery_{safe_surgery}'] = 1
                        row[f'surgery_{safe_surgery}_eye'] = surgery['eye']

                # Fill systemic conditions (BINARY)
                if include_systemic:
                    for cond in patient_systemic.get(patient['patient_id'], []):
                        safe_code = make_safe_column_name(cond['icd10_code'])
                        row[f'systemic_{safe_code}'] = 1

                # Fill ocular medications (BINARY)
                if include_medications:
                    for med in patient_ocular_meds.get(patient['patient_id'], []):
                        safe_med = make_safe_column_name(med['generic_name'])
                        row[f'ocular_med_{safe_med}'] = 1
                        row[f'ocular_med_{safe_med}_eye'] = med['eye']
                        row[f'ocular_med_{safe_med}_days'] = med['last_application_days']

                # Fill systemic medications (BINARY)
                if include_medications:
                    for med in patient_systemic_meds.get(patient['patient_id'], []):
                        safe_med = make_safe_column_name(med['generic_name'])
                        row[f'systemic_med_{safe_med}'] = 1
                        row[f'systemic_med_{safe_med}_days'] = med['last_application_days']

                # Extract and fill generic components
                if include_medications:
                    # Combine all patient medications
                    patient_all_meds = []

                    # Add ocular medications
                    for med in patient_ocular_meds.get(patient['patient_id'], []):
                        patient_all_meds.append({
                            'generic_name': med['generic_name']
                        })

                    # Add systemic medications
                    for med in patient_systemic_meds.get(patient['patient_id'], []):
                        patient_all_meds.append({
                            'generic_name': med['generic_name']
                        })

                    # Extract generic components dynamically
                    generic_flags = extract_generic_components_dynamic(patient_all_meds, all_generic_components)

                    # Add to row
                    for key, value in generic_flags.items():
                        if key in final_columns:
                            row[key] = value

                export_data.append(row)

            # ============================================================
            # STEP 6: Generate export file
            # ============================================================

            if export_format == 'csv':
                # Generate CSV
                output = io.StringIO()
                if export_data:
                    writer = csv.DictWriter(output, fieldnames=final_columns, extrasaction='ignore')
                    writer.writeheader()
                    writer.writerows(export_data)

                from flask import Response
                filename_type = 'sensitive' if data_type == 'sensitive' else 'anonymized'
                return Response(
                    output.getvalue(),
                    mimetype='text/csv',
                    headers={
                        'Content-Disposition': f'attachment; filename=raman_export_binary_{filename_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
                    }
                )

            elif export_format == 'excel':
                # Generate Excel file
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment
                    from openpyxl.utils import get_column_letter

                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Patient Data"

                    if export_data:
                        # Write headers
                        header_fill = PatternFill(start_color="3498db", end_color="3498db", fill_type="solid")
                        header_font = Font(bold=True, color="FFFFFF")

                        for col_idx, fieldname in enumerate(final_columns, 1):
                            cell = ws.cell(row=1, column=col_idx, value=fieldname)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal='center')

                        # Write data
                        for row_idx, data_row in enumerate(export_data, 2):
                            for col_idx, fieldname in enumerate(final_columns, 1):
                                value = data_row.get(fieldname, '')
                                ws.cell(row=row_idx, column=col_idx, value=value)

                        # Auto-adjust column widths
                        for col_idx in range(1, len(final_columns) + 1):
                            ws.column_dimensions[get_column_letter(col_idx)].width = 15

                    # Save to BytesIO
                    excel_output = io.BytesIO()
                    wb.save(excel_output)
                    excel_output.seek(0)

                    from flask import Response
                    filename_type = 'sensitive' if data_type == 'sensitive' else 'anonymized'
                    return Response(
                        excel_output.getvalue(),
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        headers={
                            'Content-Disposition': f'attachment; filename=raman_export_binary_{filename_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                        }
                    )

                except ImportError:
                    flash('Excel export requires openpyxl. Please run: pip install openpyxl', 'error')
                    return redirect(url_for('export_data'))

        cur.close()
        conn.close()
        return render_template('export_data.html', stats=stats)

    except Exception as e:
        flash(f'Error with export: {str(e)}', 'error')
        if conn:
            conn.close()
        # Return proper stats structure even on error
        return render_template('export_data.html',
                               stats={'total_patients': 0, 'gender': {'M': 0, 'F': 0}, 'age_distribution': []})


# Settings Routes

@app.route('/settings')
@admin_required
def settings():
    """Settings page"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return render_template('settings.html', stats={})

    try:
        cur = conn.cursor()

        # Get counts for display
        cur.execute('SELECT COUNT(*) FROM icd10_ocular_conditions WHERE active = TRUE')
        ocular_icd10 = cur.fetchone()[0]

        cur.execute('SELECT COUNT(*) FROM icd10_systemic_conditions WHERE active = TRUE')
        systemic_icd10 = cur.fetchone()[0]

        cur.execute('SELECT COUNT(*) FROM medications WHERE active = TRUE')
        medications = cur.fetchone()[0]

        cur.execute('SELECT COUNT(*) FROM surgeries WHERE active = TRUE')
        surgeries = cur.fetchone()[0]

        stats = {
            'ocular_icd10': ocular_icd10,
            'systemic_icd10': systemic_icd10,
            'medications': medications,
            'surgeries': surgeries
        }

        cur.close()
        conn.close()

        return render_template('settings.html', stats=stats)
    except Exception as e:
        flash(f'Error loading settings: {str(e)}', 'error')
        if conn:
            conn.close()
        return render_template('settings.html', stats={})


@app.route('/settings/icd10-ocular')
@admin_required
def settings_icd10_ocular():
    """Manage ICD-10 ocular conditions"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('settings'))

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute('SELECT * FROM icd10_ocular_conditions ORDER BY code')
        codes = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('settings_icd10_ocular.html', codes=codes)
    except Exception as e:
        flash(f'Error loading ICD-10 codes: {str(e)}', 'error')
        if conn:
            conn.close()
        return redirect(url_for('settings'))


@app.route('/settings/icd10-systemic')
@admin_required
def settings_icd10_systemic():
    """Manage ICD-10 systemic conditions"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('settings'))

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute('SELECT * FROM icd10_systemic_conditions ORDER BY code')
        codes = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('settings_icd10_systemic.html', codes=codes)
    except Exception as e:
        flash(f'Error loading ICD-10 codes: {str(e)}', 'error')
        if conn:
            conn.close()
        return redirect(url_for('settings'))


@app.route('/settings/medications')
@admin_required
def settings_medications():
    """Manage medications"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('settings'))

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute('SELECT * FROM medications ORDER BY trade_name')
        medications = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('settings_medications.html', medications=medications)
    except Exception as e:
        flash(f'Error loading medications: {str(e)}', 'error')
        if conn:
            conn.close()
        return redirect(url_for('settings'))


@app.route('/settings/surgeries')
@admin_required
def settings_surgeries():
    """Manage surgeries"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('settings'))

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)
        cur.execute('SELECT * FROM surgeries ORDER BY code')
        surgeries = cur.fetchall()
        cur.close()
        conn.close()
        return render_template('settings_surgeries.html', surgeries=surgeries)
    except Exception as e:
        flash(f'Error loading surgeries: {str(e)}', 'error')
        if conn:
            conn.close()
        return redirect(url_for('settings'))


# =============== BACKUP ROUTES ===============
@app.route('/settings/backup')
@admin_required
def settings_backup():
    """Backup and restore management page"""
    config = load_backup_config()

    # Get list of existing backups
    backup_dir = config['backup_dir']
    backups = []

    if os.path.exists(backup_dir):
        for file in sorted(os.listdir(backup_dir), reverse=True):
            if file.endswith('.sql') or file.endswith('.dump'):
                file_path = os.path.join(backup_dir, file)
                file_stat = os.stat(file_path)
                backups.append({
                    'filename': file,
                    'size': format_file_size(file_stat.st_size),
                    'created': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S'),
                    'timestamp': file_stat.st_mtime
                })

    return render_template('settings_backup.html',
                           config=config,
                           backups=backups[:20])  # Show last 20 backups


@app.route('/api/browse-directory', methods=['POST'])
@admin_required
def browse_directory():
    """Browse server directory structure with extended support for external drives"""
    current_path = request.json.get('path', '/')

    # Normalize path
    current_path = os.path.normpath(current_path)

    # Extended allowed roots for Raspberry Pi and external drives
    allowed_roots = [
        '/backups',  # Default backup location
        '/mnt',  # Standard mount point for external drives
        '/media',  # Alternative mount point (Raspberry Pi often uses this)
        '/external_backup',  # Docker mapped external drive
        '/home',  # User home directories
        '/var/backups',  # System backup location
        '/tmp'  # Temporary directory
    ]

    # Dynamically add specific mount points if they exist (for Raspberry Pi)
    if os.path.exists('/mnt/medical_backups'):
        allowed_roots.append('/mnt/medical_backups')

    # For Unraid - add common paths
    if os.path.exists('/mnt/user'):
        allowed_roots.append('/mnt/user')
    if os.path.exists('/mnt/disks'):
        allowed_roots.append('/mnt/disks')

    # Check if path is within allowed roots
    path_allowed = False
    for root in allowed_roots:
        if current_path.startswith(root) or current_path == '/':
            path_allowed = True
            break

    if not path_allowed:
        return jsonify({'error': 'Access to this directory is not allowed'}), 403

    try:
        entries = []

        # Add parent directory option if not at root
        if current_path != '/':
            entries.append({
                'name': '..',
                'path': os.path.dirname(current_path),
                'type': 'directory'
            })

        # List directory contents
        if current_path == '/':
            # Show only allowed root directories that exist
            for root in allowed_roots:
                if os.path.exists(root):
                    entries.append({
                        'name': root,
                        'path': root,
                        'type': 'directory'
                    })
        else:
            # List actual directory contents
            try:
                for entry in sorted(os.listdir(current_path)):
                    entry_path = os.path.join(current_path, entry)

                    # Skip hidden files/directories
                    if entry.startswith('.'):
                        continue

                    try:
                        if os.path.isdir(entry_path):
                            # Check if it's a mount point (useful for external drives)
                            is_mount = os.path.ismount(entry_path)

                            # Get directory size if it's a mount point
                            mount_info = ""
                            if is_mount:
                                try:
                                    statvfs = os.statvfs(entry_path)
                                    free_space = statvfs.f_frsize * statvfs.f_bavail
                                    total_space = statvfs.f_frsize * statvfs.f_blocks
                                    free_gb = free_space / (1024 ** 3)
                                    total_gb = total_space / (1024 ** 3)
                                    mount_info = f" ({free_gb:.1f}GB free / {total_gb:.1f}GB total)"
                                except:
                                    mount_info = " (mount point)"

                            entries.append({
                                'name': entry + mount_info,
                                'path': entry_path,
                                'type': 'directory',
                                'is_mount': is_mount
                            })
                    except PermissionError:
                        # Skip directories we don't have permission to access
                        continue
            except PermissionError:
                # If we can't list the directory at all
                return jsonify({
                    'error': 'Permission denied for this directory',
                    'current_path': current_path,
                    'entries': []
                }), 403

        return jsonify({
            'current_path': current_path,
            'entries': entries
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500


def check_backup_location(backup_dir):
    """Check if backup location is available and writable"""
    try:
        # Check if directory exists
        if not os.path.exists(backup_dir):
            os.makedirs(backup_dir, exist_ok=True)

        # Check if it's writable
        test_file = os.path.join(backup_dir, '.write_test')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)

        # Check if it's an external drive (optional)
        mount_point = os.path.ismount(backup_dir) or os.path.ismount(os.path.dirname(backup_dir))

        # Check available space
        statvfs = os.statvfs(backup_dir)
        free_space = statvfs.f_frsize * statvfs.f_bavail
        free_space_gb = free_space / (1024 ** 3)

        return {
            'available': True,
            'writable': True,
            'is_external': mount_point,
            'free_space_gb': round(free_space_gb, 2)
        }
    except Exception as e:
        return {
            'available': False,
            'error': str(e)
        }


@app.route('/api/backup-drive-status')
@admin_required
def backup_drive_status():
    """Check the status of the backup location (especially for external drives)"""
    config = load_backup_config()
    backup_dir = config.get('backup_dir', '/backups')

    try:
        # Check if directory exists
        if not os.path.exists(backup_dir):
            return jsonify({
                'available': False,
                'error': 'Backup directory does not exist'
            })

        # Check if it's writable
        writable = False
        try:
            test_file = os.path.join(backup_dir, '.write_test')
            with open(test_file, 'w') as f:
                f.write('test')
            os.remove(test_file)
            writable = True
        except:
            writable = False

        # Check if it's a mount point (indicates external drive)
        is_external = os.path.ismount(backup_dir) or os.path.ismount(os.path.dirname(backup_dir))

        # Get mount point path
        mount_point = backup_dir
        if not os.path.ismount(backup_dir):
            # Find the actual mount point
            path = backup_dir
            while path != '/':
                if os.path.ismount(path):
                    mount_point = path
                    break
                path = os.path.dirname(path)

        # Check available space
        statvfs = os.statvfs(backup_dir)
        free_space = statvfs.f_frsize * statvfs.f_bavail
        total_space = statvfs.f_frsize * statvfs.f_blocks
        free_space_gb = free_space / (1024 ** 3)
        total_space_gb = total_space / (1024 ** 3)
        used_percent = ((total_space - free_space) / total_space) * 100 if total_space > 0 else 0

        return jsonify({
            'available': True,
            'writable': writable,
            'is_external': is_external,
            'mount_point': mount_point,
            'backup_dir': backup_dir,
            'free_space_gb': round(free_space_gb, 2),
            'total_space_gb': round(total_space_gb, 2),
            'used_percent': round(used_percent, 1)
        })

    except Exception as e:
        return jsonify({
            'available': False,
            'error': str(e)
        })


@app.route('/api/create-backup', methods=['POST'])
@admin_required
def create_backup():
    """Create a database backup with improved error handling"""
    config = load_backup_config()
    backup_dir = config['backup_dir']

    # Create backup directory if it doesn't exist
    os.makedirs(backup_dir, exist_ok=True)

    # Check if backup location is available
    location_status = check_backup_location(backup_dir)
    if not location_status['available']:
        return jsonify({
            'error': f"Backup location unavailable: {location_status.get('error', 'Unknown error')}"
        }), 500

    # Warn if low space (less than 1GB)
    if location_status.get('free_space_gb', 0) < 1:
        return jsonify({
            'error': f"Insufficient space: only {location_status['free_space_gb']}GB available"
        }), 500

    # Generate backup filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_file = os.path.join(backup_dir, f'raman_backup_{timestamp}.sql')

    try:
        # Create pg_dump command
        env = os.environ.copy()
        env['PGPASSWORD'] = DB_CONFIG['password']

        # Use --no-owner and --no-acl for better compatibility
        cmd = [
            'pg_dump',
            '-h', DB_CONFIG['host'],
            '-p', DB_CONFIG['port'],
            '-U', DB_CONFIG['user'],
            '-d', DB_CONFIG['dbname'],
            '-f', backup_file,
            '--no-owner',  # Don't output ownership commands
            '--no-acl',  # Don't output access privileges
            '--clean',  # Add DROP statements before CREATE
            '--if-exists',  # Use IF EXISTS for DROP statements
            '--no-password',
            '--verbose'
        ]

        # Execute backup
        print(f"Executing backup command: {' '.join(cmd)}")
        result = subprocess.run(cmd, env=env, capture_output=True, text=True, timeout=60)

        # Check if file was created even if there were warnings
        if os.path.exists(backup_file) and os.path.getsize(backup_file) > 0:
            file_size = os.path.getsize(backup_file)

            # Log any warnings but still consider it successful if file exists
            if result.stderr:
                print(f"Backup completed with warnings: {result.stderr[:500]}")  # Log first 500 chars

            return jsonify({
                'success': True,
                'message': f'Backup created successfully',
                'filename': os.path.basename(backup_file),
                'size': format_file_size(file_size),
                'warnings': 'Check server logs for details' if result.stderr else None
            })
        else:
            # Real error - no file created
            error_msg = result.stderr if result.stderr else 'Unknown error - backup file not created'
            print(f"Backup failed: {error_msg}")

            # Clean up empty file if it exists
            if os.path.exists(backup_file):
                os.remove(backup_file)

            return jsonify({
                'error': f'Backup failed: {error_msg[:500]}'  # Limit error message length
            }), 500

    except subprocess.TimeoutExpired:
        return jsonify({
            'error': 'Backup timed out after 60 seconds. Database might be too large or busy.'
        }), 500

    except FileNotFoundError:
        return jsonify({
            'error': 'pg_dump not found. Please ensure PostgreSQL client tools are installed.'
        }), 500

    except Exception as e:
        print(f"Backup error: {str(e)}")
        return jsonify({'error': f'Backup failed: {str(e)[:500]}'}), 500


# Alternative backup method using SQL COPY commands if pg_dump fails
@app.route('/api/create-backup-sql', methods=['POST'])
@admin_required
def create_backup_sql():
    """Alternative backup using SQL COPY commands - exports data as CSV"""
    config = load_backup_config()
    backup_dir = config['backup_dir']

    # Create backup directory
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_folder = os.path.join(backup_dir, f'backup_{timestamp}')
    os.makedirs(backup_folder, exist_ok=True)

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        cur = conn.cursor()

        # Get list of all tables
        cur.execute("""
            SELECT tablename 
            FROM pg_tables 
            WHERE schemaname = 'public' 
            ORDER BY tablename
        """)
        tables = [row[0] for row in cur.fetchall()]

        exported_tables = []

        # Export each table to CSV
        for table in tables:
            csv_file = os.path.join(backup_folder, f'{table}.csv')

            # Get table data
            cur.execute(f"SELECT * FROM {table}")
            rows = cur.fetchall()

            # Get column names
            col_names = [desc[0] for desc in cur.description]

            # Write to CSV
            import csv
            with open(csv_file, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(col_names)
                writer.writerows(rows)

            exported_tables.append(table)

        # Create metadata file
        metadata = {
            'timestamp': timestamp,
            'tables': exported_tables,
            'database': DB_CONFIG['dbname'],
            'backup_type': 'csv_export'
        }

        metadata_file = os.path.join(backup_folder, 'backup_metadata.json')
        with open(metadata_file, 'w') as f:
            json.dump(metadata, f, indent=2)

        # Calculate total size
        total_size = sum(
            os.path.getsize(os.path.join(backup_folder, f))
            for f in os.listdir(backup_folder)
        )

        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'CSV backup created successfully',
            'folder': os.path.basename(backup_folder),
            'tables_exported': len(exported_tables),
            'size': format_file_size(total_size)
        })

    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'error': f'Backup failed: {str(e)}'}), 500


# Diagnostic function to check pg_dump version
@app.route('/api/check-backup-tools')
@admin_required
def check_backup_tools():
    """Check if backup tools are properly installed"""
    diagnostics = {}

    # Check pg_dump
    try:
        result = subprocess.run(['pg_dump', '--version'], capture_output=True, text=True, timeout=5)
        diagnostics['pg_dump'] = {
            'installed': True,
            'version': result.stdout.strip() if result.stdout else 'Unknown'
        }
    except FileNotFoundError:
        diagnostics['pg_dump'] = {
            'installed': False,
            'error': 'pg_dump not found in PATH'
        }
    except Exception as e:
        diagnostics['pg_dump'] = {
            'installed': False,
            'error': str(e)
        }

    # Check PostgreSQL server version
    try:
        conn = get_db_connection()
        if conn:
            cur = conn.cursor()
            cur.execute("SELECT version()")
            diagnostics['postgresql'] = {
                'connected': True,
                'version': cur.fetchone()[0]
            }
            cur.close()
            conn.close()
        else:
            diagnostics['postgresql'] = {
                'connected': False,
                'error': 'Could not connect to database'
            }
    except Exception as e:
        diagnostics['postgresql'] = {
            'connected': False,
            'error': str(e)
        }

    # Check backup directory
    config = load_backup_config()
    backup_dir = config['backup_dir']

    try:
        os.makedirs(backup_dir, exist_ok=True)
        test_file = os.path.join(backup_dir, '.test_write')
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)

        diagnostics['backup_directory'] = {
            'path': backup_dir,
            'writable': True,
            'exists': True
        }
    except Exception as e:
        diagnostics['backup_directory'] = {
            'path': backup_dir,
            'writable': False,
            'error': str(e)
        }

    return jsonify(diagnostics)


@app.route('/api/restore-backup', methods=['POST'])
@admin_required
def restore_backup():
    """Restore a database backup"""
    backup_filename = request.json.get('filename')
    if not backup_filename:
        return jsonify({'error': 'No backup file specified'}), 400

    config = load_backup_config()
    backup_file = os.path.join(config['backup_dir'], backup_filename)

    if not os.path.exists(backup_file):
        return jsonify({'error': 'Backup file not found'}), 404

    try:
        # Set environment for psql
        env = os.environ.copy()
        env['PGPASSWORD'] = DB_CONFIG['password']

        # Drop existing connections to database
        drop_connections_cmd = [
            'psql',
            '-h', DB_CONFIG['host'],
            '-p', DB_CONFIG['port'],
            '-U', DB_CONFIG['user'],
            '-d', 'postgres',
            '--no-password',
            '-c',
            f"SELECT pg_terminate_backend(pid) FROM pg_stat_activity WHERE datname = '{DB_CONFIG['dbname']}' AND pid <> pg_backend_pid();"
        ]

        subprocess.run(drop_connections_cmd, env=env, capture_output=True)

        # Restore the backup
        cmd = [
            'psql',
            '-h', DB_CONFIG['host'],
            '-p', DB_CONFIG['port'],
            '-U', DB_CONFIG['user'],
            '-d', DB_CONFIG['dbname'],
            '-f', backup_file,
            '--no-password'
        ]

        result = subprocess.run(cmd, env=env, capture_output=True, text=True)

        if result.returncode == 0:
            return jsonify({
                'success': True,
                'message': 'Database restored successfully. Please log in again.'
            })
        else:
            return jsonify({
                'error': f'Restore failed: {result.stderr}'
            }), 500

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/delete-backup', methods=['POST'])
@admin_required
def delete_backup():
    """Delete a backup file"""
    backup_filename = request.json.get('filename')
    if not backup_filename:
        return jsonify({'error': 'No backup file specified'}), 400

    config = load_backup_config()
    backup_file = os.path.join(config['backup_dir'], backup_filename)

    if not os.path.exists(backup_file):
        return jsonify({'error': 'Backup file not found'}), 404

    try:
        os.remove(backup_file)
        return jsonify({
            'success': True,
            'message': 'Backup deleted successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/backup-config', methods=['POST'])
@admin_required
def update_backup_config():
    """Update backup configuration"""
    data = request.json
    config = load_backup_config()

    # Update configuration
    if 'backup_dir' in data:
        # Validate directory
        backup_dir = data['backup_dir']
        try:
            os.makedirs(backup_dir, exist_ok=True)
            config['backup_dir'] = backup_dir
        except Exception as e:
            return jsonify({'error': f'Invalid directory: {e}'}), 400

    if 'schedule' in data:
        config['schedule'] = data['schedule']

    if 'retention_days' in data:
        config['retention_days'] = int(data['retention_days'])

    if 'auto_backup_enabled' in data:
        config['auto_backup_enabled'] = data['auto_backup_enabled']

    save_backup_config(config)

    # Restart scheduler if needed
    if config['auto_backup_enabled']:
        start_backup_scheduler(config)
    else:
        stop_backup_scheduler()

    return jsonify({
        'success': True,
        'message': 'Configuration updated successfully'
    })


@app.route('/download-backup/<filename>')
@admin_required
def download_backup(filename):
    """Download a backup file"""
    from flask import send_file
    config = load_backup_config()
    backup_file = os.path.join(config['backup_dir'], filename)

    if not os.path.exists(backup_file):
        flash('Backup file not found', 'error')
        return redirect(url_for('settings_backup'))

    return send_file(backup_file, as_attachment=True, download_name=filename)


# User Management Routes

@app.route('/user-management')
@admin_required
def user_management():
    """Manage users"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return render_template('user_management.html', users=[], stats={})

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        # Get all users
        cur.execute('SELECT * FROM users ORDER BY user_id')
        users = cur.fetchall()

        # Get statistics
        cur.execute('SELECT COUNT(*) as total FROM users')
        total_users = cur.fetchone()['total']

        cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'Administrator'")
        administrators = cur.fetchone()['count']

        cur.execute("SELECT COUNT(*) as count FROM users WHERE role = 'Staff'")
        staff = cur.fetchone()['count']

        stats = {
            'total_users': total_users,
            'administrators': administrators,
            'staff': staff
        }

        cur.close()
        conn.close()

        return render_template('user_management.html', users=users, stats=stats)
    except Exception as e:
        flash(f'Error loading users: {str(e)}', 'error')
        if conn:
            conn.close()
        return render_template('user_management.html', users=[], stats={})


@app.route('/create-user', methods=['POST'])
@admin_required
def create_user():
    """Create new user"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('user_management'))

    try:
        username = request.form.get('username')
        email = request.form.get('email')
        password = request.form.get('password')
        role = request.form.get('role')

        password_hash = bcrypt.generate_password_hash(password).decode('utf-8')

        cur = conn.cursor()
        cur.execute('''
            INSERT INTO users (username, password_hash, email, role)
            VALUES (%s, %s, %s, %s)
        ''', (username, password_hash, email, role))
        conn.commit()
        cur.close()
        conn.close()

        flash(f'User {username} created successfully!', 'success')
    except Exception as e:
        flash(f'Error creating user: {str(e)}', 'error')
        if conn:
            conn.rollback()
            conn.close()

    return redirect(url_for('user_management'))


@app.route('/update-user', methods=['POST'])
@admin_required
def update_user():
    """Update user"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('user_management'))

    try:
        user_id = request.form.get('user_id')
        username = request.form.get('username')
        email = request.form.get('email')
        role = request.form.get('role')
        password = request.form.get('password')

        cur = conn.cursor()

        if password:
            password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
            cur.execute('''
                UPDATE users
                SET username = %s, email = %s, role = %s, password_hash = %s
                WHERE user_id = %s
            ''', (username, email, role, password_hash, user_id))
        else:
            cur.execute('''
                UPDATE users
                SET username = %s, email = %s, role = %s
                WHERE user_id = %s
            ''', (username, email, role, user_id))

        conn.commit()
        cur.close()
        conn.close()

        flash(f'User {username} updated successfully!', 'success')
    except Exception as e:
        flash(f'Error updating user: {str(e)}', 'error')
        if conn:
            conn.rollback()
            conn.close()

    return redirect(url_for('user_management'))


@app.route('/delete-user', methods=['POST'])
@admin_required
def delete_user():
    """Delete user"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('user_management'))

    try:
        user_id = request.form.get('user_id')

        cur = conn.cursor()
        cur.execute('DELETE FROM users WHERE user_id = %s', (user_id,))
        conn.commit()
        cur.close()
        conn.close()

        flash('User deleted successfully!', 'success')
    except Exception as e:
        flash(f'Error deleting user: {str(e)}', 'error')
        if conn:
            conn.rollback()
            conn.close()

    return redirect(url_for('user_management'))


@app.route('/reset-user-password', methods=['POST'])
@admin_required
def reset_user_password():
    """Reset user password to default"""
    conn = get_db_connection()
    if not conn:
        flash('Database connection error', 'error')
        return redirect(url_for('user_management'))

    try:
        user_id = request.form.get('user_id')
        default_password = 'password123'
        password_hash = bcrypt.generate_password_hash(default_password).decode('utf-8')

        cur = conn.cursor()
        cur.execute('UPDATE users SET password_hash = %s WHERE user_id = %s', (password_hash, user_id))
        conn.commit()
        cur.close()
        conn.close()

        flash(f'Password reset to: {default_password}', 'success')
    except Exception as e:
        flash(f'Error resetting password: {str(e)}', 'error')
        if conn:
            conn.rollback()
            conn.close()

    return redirect(url_for('user_management'))


# API Routes for Managing Reference Data

@app.route('/api/icd10-ocular', methods=['GET', 'POST', 'PUT', 'DELETE'])
@admin_required
def api_icd10_ocular():
    """API for managing ICD-10 ocular conditions"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        if request.method == 'GET':
            cur.execute('SELECT * FROM icd10_ocular_conditions ORDER BY code')
            codes = cur.fetchall()
            cur.close()
            conn.close()
            return jsonify(codes)

        elif request.method == 'POST':
            data = request.get_json()
            cur.execute('''
                INSERT INTO icd10_ocular_conditions (code, description, category, active)
                VALUES (%s, %s, %s, %s)
                RETURNING *
            ''', (data['code'], data['description'], data.get('category', ''), data.get('active', True)))
            new_code = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(new_code), 201

        elif request.method == 'PUT':
            data = request.get_json()
            cur.execute('''
                UPDATE icd10_ocular_conditions
                SET description = %s, category = %s, active = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING *
            ''', (data['description'], data.get('category', ''), data.get('active', True), data['id']))
            updated_code = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(updated_code)

        elif request.method == 'DELETE':
            data = request.get_json()

            # Check for permanent delete flag
            if data.get('permanent', False):
                # Permanently delete the record
                cur.execute('DELETE FROM icd10_ocular_conditions WHERE id = %s', (data['id'],))
            else:
                # Just deactivate (default behavior)
                cur.execute('''
                    UPDATE icd10_ocular_conditions
                    SET active = FALSE, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                ''', (data['id'],))

            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': True})

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/icd10-systemic', methods=['GET', 'POST', 'PUT', 'DELETE'])
@admin_required
def api_icd10_systemic():
    """API for managing ICD-10 systemic conditions"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        if request.method == 'GET':
            cur.execute('SELECT * FROM icd10_systemic_conditions ORDER BY code')
            codes = cur.fetchall()
            cur.close()
            conn.close()
            return jsonify(codes)

        elif request.method == 'POST':
            data = request.get_json()
            cur.execute('''
                INSERT INTO icd10_systemic_conditions (code, description, category, active)
                VALUES (%s, %s, %s, %s)
                RETURNING *
            ''', (data['code'], data['description'], data.get('category', ''), data.get('active', True)))
            new_code = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(new_code), 201

        elif request.method == 'PUT':
            data = request.get_json()
            cur.execute('''
                UPDATE icd10_systemic_conditions
                SET description = %s, category = %s, active = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING *
            ''', (data['description'], data.get('category', ''), data.get('active', True), data['id']))
            updated_code = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(updated_code)

        elif request.method == 'DELETE':
            data = request.get_json()

            # Check for permanent delete flag
            if data.get('permanent', False):
                # Permanently delete the record
                cur.execute('DELETE FROM icd10_systemic_conditions WHERE id = %s', (data['id'],))
            else:
                # Just deactivate (default behavior)
                cur.execute('''
                    UPDATE icd10_systemic_conditions
                    SET active = FALSE, updated_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                ''', (data['id'],))

            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': True})

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': str(e)}), 500


# Bulk ICD-10 Upload Routes and Functions
@app.route('/settings/icd10-bulk-upload/<code_type>')
@admin_required
def icd10_bulk_upload(code_type):
    """Bulk upload page for ICD-10 codes"""
    if code_type not in ['ocular', 'systemic']:
        flash('Invalid code type', 'error')
        return redirect(url_for('settings'))

    return render_template('icd10_bulk_upload.html', code_type=code_type)


@app.route('/api/icd10-bulk-preview', methods=['POST'])
@admin_required
def icd10_bulk_preview():
    """Preview CSV/Excel data before importing"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    # Check file extension
    filename = file.filename.lower()
    if not any(filename.endswith(ext) for ext in ['.csv', '.xls', '.xlsx']):
        return jsonify({'error': 'Invalid file format. Please upload CSV, XLS, or XLSX'}), 400

    try:
        import pandas as pd

        # Read file based on extension
        if filename.endswith('.csv'):
            # Try different encodings
            try:
                df = pd.read_csv(file, encoding='utf-8')
            except UnicodeDecodeError:
                file.seek(0)
                df = pd.read_csv(file, encoding='latin-1')
        else:
            df = pd.read_excel(file)

        # Get first 10 rows as preview
        preview_data = df.head(10).fillna('').to_dict('records')

        # Get column names
        columns = df.columns.tolist()

        # Try to auto-detect code and description columns
        auto_mapping = {
            'code_column': None,
            'description_column': None,
            'category_column': None
        }

        for col in columns:
            col_lower = col.lower()
            # Auto-detect code column
            if any(term in col_lower for term in ['code', 'icd', 'id', 'codigo']):
                if not auto_mapping['code_column']:
                    auto_mapping['code_column'] = col
            # Auto-detect description column
            elif any(term in col_lower for term in ['description', 'desc', 'name', 'text', 'descripcion', 'nombre']):
                if not auto_mapping['description_column']:
                    auto_mapping['description_column'] = col
            # Auto-detect category column
            elif any(term in col_lower for term in ['category', 'cat', 'type', 'group', 'categoria', 'tipo']):
                if not auto_mapping['category_column']:
                    auto_mapping['category_column'] = col

        return jsonify({
            'success': True,
            'columns': columns,
            'preview': preview_data,
            'total_rows': len(df),
            'auto_mapping': auto_mapping
        })

    except Exception as e:
        return jsonify({'error': f'Error reading file: {str(e)}'}), 500


@app.route('/api/icd10-bulk-import', methods=['POST'])
@admin_required
def icd10_bulk_import():
    """Import ICD-10 codes from CSV/Excel with field mapping"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    code_type = request.form.get('code_type')  # 'ocular' or 'systemic'
    code_column = request.form.get('code_column')
    description_column = request.form.get('description_column')
    category_column = request.form.get('category_column')  # Optional

    if not all([code_type, code_column, description_column]):
        return jsonify({'error': 'Missing required field mappings'}), 400

    if code_type not in ['ocular', 'systemic']:
        return jsonify({'error': 'Invalid code type'}), 400

    table_name = f'icd10_{code_type}_conditions'

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        import pandas as pd

        # Read file
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            try:
                df = pd.read_csv(file, encoding='utf-8')
            except UnicodeDecodeError:
                file.seek(0)
                df = pd.read_csv(file, encoding='latin-1')
        else:
            df = pd.read_excel(file)

        cur = conn.cursor()

        imported = 0
        skipped = 0
        errors = []

        for index, row in df.iterrows():
            try:
                code = str(row[code_column]).strip() if pd.notna(row[code_column]) else None
                description = str(row[description_column]).strip() if pd.notna(row[description_column]) else None

                if not code or not description:
                    skipped += 1
                    continue

                # Get category if column is mapped
                category = None
                if category_column and category_column in row:
                    category = str(row[category_column]).strip() if pd.notna(row[category_column]) else None

                # Auto-detect category for ocular codes if not provided
                if code_type == 'ocular' and not category:
                    if code.startswith('H0'):
                        category = 'Eyelid and orbit'
                    elif code.startswith('H1'):
                        category = 'Conjunctiva and cornea'
                    elif code.startswith('H2'):
                        category = 'Lens'
                    elif code.startswith('H3'):
                        category = 'Retina and choroid'
                    elif code.startswith('H4'):
                        category = 'Glaucoma'
                    elif code.startswith('H5'):
                        category = 'Visual disorders'

                # Auto-detect category for systemic codes if not provided
                elif code_type == 'systemic' and not category:
                    first_char = code[0] if code else ''
                    if first_char in 'AB':
                        category = 'Infectious diseases'
                    elif first_char == 'C':
                        category = 'Neoplasms'
                    elif first_char == 'D':
                        category = 'Blood diseases'
                    elif first_char == 'E':
                        category = 'Endocrine'
                    elif first_char == 'F':
                        category = 'Mental disorders'
                    elif first_char == 'G':
                        category = 'Nervous system'
                    elif first_char == 'I':
                        category = 'Circulatory'
                    elif first_char == 'J':
                        category = 'Respiratory'
                    elif first_char == 'K':
                        category = 'Digestive'
                    elif first_char == 'L':
                        category = 'Skin'
                    elif first_char == 'M':
                        category = 'Musculoskeletal'
                    elif first_char == 'N':
                        category = 'Genitourinary'
                    elif first_char == 'R':
                        category = 'Symptoms'
                    elif first_char == 'Z':
                        category = 'Health factors'

                # Insert or update
                cur.execute(f'''
                    INSERT INTO {table_name} (code, description, category, active)
                    VALUES (%s, %s, %s, TRUE)
                    ON CONFLICT (code) DO UPDATE
                    SET description = EXCLUDED.description,
                        category = EXCLUDED.category,
                        active = TRUE,
                        updated_at = CURRENT_TIMESTAMP
                ''', (code, description, category))

                imported += 1

            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                skipped += 1

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:10] if errors else []  # Return first 10 errors
        })

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': f'Import failed: {str(e)}'}), 500


@app.route('/api/icd10-export/<code_type>')
@admin_required
def icd10_export(code_type):
    """Export ICD-10 codes to Excel"""
    if code_type not in ['ocular', 'systemic']:
        return jsonify({'error': 'Invalid code type'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        import pandas as pd
        from io import BytesIO

        table_name = f'icd10_{code_type}_conditions'

        # Get all codes
        query = f'''
            SELECT code as "ICD-10 Code", 
                   description as "Description", 
                   category as "Category",
                   CASE WHEN active THEN 'Active' ELSE 'Inactive' END as "Status"
            FROM {table_name}
            ORDER BY code
        '''

        df = pd.read_sql_query(query, conn)
        conn.close()

        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=f'{code_type.title()} Codes', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets[f'{code_type.title()} Codes']
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=icd10_{code_type}_codes_{datetime.now().strftime("%Y%m%d")}.xlsx'
            }
        )

    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'error': f'Export failed: {str(e)}'}), 500


@app.route('/api/medications', methods=['GET', 'POST', 'PUT', 'DELETE'])
@admin_required
def api_medications():
    """API for managing medications"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        if request.method == 'GET':
            cur.execute('SELECT * FROM medications ORDER BY trade_name')
            medications = cur.fetchall()
            cur.close()
            conn.close()
            return jsonify(medications)

        elif request.method == 'POST':
            data = request.get_json()
            cur.execute('''
                INSERT INTO medications (trade_name, generic_name, medication_type, active)
                VALUES (%s, %s, %s, %s)
                RETURNING *
            ''', (
                data['trade_name'], data['generic_name'], data.get('medication_type', 'Both'),
                data.get('active', True)))
            new_medication = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(new_medication), 201

        elif request.method == 'PUT':
            data = request.get_json()
            cur.execute('''
                UPDATE medications
                SET trade_name = %s, generic_name = %s, medication_type = %s, 
                    active = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING *
            ''', (data['trade_name'], data['generic_name'], data.get('medication_type', 'Both'),
                  data.get('active', True), data['id']))
            updated_medication = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(updated_medication)

        elif request.method == 'DELETE':
            data = request.get_json()
            cur.execute('''
                UPDATE medications
                SET active = FALSE, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (data['id'],))
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': True})

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/surgeries', methods=['GET', 'POST', 'PUT', 'DELETE'])
@admin_required
def api_surgeries():
    """API for managing surgeries"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        cur = conn.cursor(cursor_factory=RealDictCursor)

        if request.method == 'GET':
            cur.execute('SELECT * FROM surgeries ORDER BY code')
            surgeries = cur.fetchall()
            cur.close()
            conn.close()
            return jsonify(surgeries)

        elif request.method == 'POST':
            data = request.get_json()
            cur.execute('''
                INSERT INTO surgeries (code, description, category, active)
                VALUES (%s, %s, %s, %s)
                RETURNING *
            ''', (data['code'], data['description'], data.get('category', ''), data.get('active', True)))
            new_surgery = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(new_surgery), 201

        elif request.method == 'PUT':
            data = request.get_json()
            cur.execute('''
                UPDATE surgeries
                SET description = %s, category = %s, active = %s, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
                RETURNING *
            ''', (data['description'], data.get('category', ''), data.get('active', True), data['id']))
            updated_surgery = cur.fetchone()
            conn.commit()
            cur.close()
            conn.close()
            return jsonify(updated_surgery)

        elif request.method == 'DELETE':
            data = request.get_json()
            cur.execute('''
                UPDATE surgeries
                SET active = FALSE, updated_at = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (data['id'],))
            conn.commit()
            cur.close()
            conn.close()
            return jsonify({'success': True})

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': str(e)}), 500


@app.route('/api/check_patient_id/<int:patient_id>')
@login_required
def check_patient_id(patient_id):
    """Check if a patient ID is available"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error', 'available': False}), 500

    try:
        cur = conn.cursor()
        # Check if the patient_id exists in the sensitive table
        cur.execute("SELECT patient_id FROM patients_sensitive WHERE patient_id = %s", (patient_id,))
        exists = cur.fetchone() is not None
        cur.close()
        conn.close()
        return jsonify({'available': not exists, 'patient_id': patient_id})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'error': str(e), 'available': False}), 500


@app.route('/api/next_available_patient_id')
@login_required
def api_next_available_patient_id():
    """API endpoint to get the next available patient ID"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error', 'next_id': STARTING_PATIENT_ID}), 500

    try:
        cur = conn.cursor()
        # Get the maximum patient_id from the database
        cur.execute("SELECT MAX(patient_id) FROM patients_sensitive")
        result = cur.fetchone()
        max_id = result[0] if result[0] is not None else (STARTING_PATIENT_ID - 1)

        # Next ID is max + 1, but ensure it's at least STARTING_PATIENT_ID
        next_id = max(max_id + 1, STARTING_PATIENT_ID)

        cur.close()
        conn.close()
        return jsonify({'next_id': next_id})
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'error': str(e), 'next_id': STARTING_PATIENT_ID}), 500


@app.route('/health')
def health_check():
    """Health check endpoint"""
    try:
        # Test database connection
        conn = get_db_connection()
        if conn:
            conn.close()
            return {'status': 'healthy'}, 200
        return {'status': 'unhealthy'}, 503
    except Exception as e:
        return {'status': 'unhealthy', 'error': str(e)}, 503


@app.route('/settings/medications-bulk-upload')
@admin_required
def medications_bulk_upload():
    """Bulk upload page for medications"""
    return render_template('medications_bulk_upload.html')


@app.route('/api/medications-bulk-preview', methods=['POST'])
@admin_required
def medications_bulk_preview():
    """Preview medication data before importing"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    # Check file extension
    filename = file.filename.lower()
    if not any(filename.endswith(ext) for ext in ['.csv', '.xls', '.xlsx']):
        return jsonify({'error': 'Invalid file format. Please upload CSV, XLS, or XLSX'}), 400

    try:
        import pandas as pd

        # Read file based on extension
        if filename.endswith('.csv'):
            try:
                df = pd.read_csv(file, encoding='utf-8')
            except UnicodeDecodeError:
                file.seek(0)
                df = pd.read_csv(file, encoding='latin-1')
        else:
            df = pd.read_excel(file)

        # Clean column names
        df.columns = df.columns.str.strip()

        # Get first 10 rows as preview
        preview_data = df.head(10).fillna('').to_dict('records')

        # Get column names
        columns = df.columns.tolist()

        # Try to auto-detect columns
        auto_mapping = {
            'trade_column': None,
            'generic_column': None,
            'type_column': None
        }

        for col in columns:
            col_lower = col.lower()
            # Auto-detect trade name column
            if any(term in col_lower for term in ['trade', 'brand', 'product']):
                if not auto_mapping['trade_column']:
                    auto_mapping['trade_column'] = col
            # Auto-detect generic name column
            elif any(term in col_lower for term in ['generic', 'substance', 'active', 'ingredient']):
                if not auto_mapping['generic_column']:
                    auto_mapping['generic_column'] = col
            # Auto-detect type column
            elif any(term in col_lower for term in ['type', 'category', 'class']):
                if not auto_mapping['type_column']:
                    auto_mapping['type_column'] = col

        return jsonify({
            'success': True,
            'columns': columns,
            'preview': preview_data,
            'total_rows': len(df),
            'auto_mapping': auto_mapping
        })

    except Exception as e:
        return jsonify({'error': f'Error reading file: {str(e)}'}), 500


@app.route('/api/medications-bulk-import', methods=['POST'])
@admin_required
def medications_bulk_import():
    """Import medications from CSV/Excel with field mapping"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    trade_column = request.form.get('trade_column')
    generic_column = request.form.get('generic_column')
    type_column = request.form.get('type_column')  # Optional

    if not all([trade_column, generic_column]):
        return jsonify({'error': 'Missing required field mappings'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        import pandas as pd

        # Read file
        filename = file.filename.lower()
        if filename.endswith('.csv'):
            try:
                df = pd.read_csv(file, encoding='utf-8')
            except UnicodeDecodeError:
                file.seek(0)
                df = pd.read_csv(file, encoding='latin-1')
        else:
            df = pd.read_excel(file)

        # Clean column names
        df.columns = df.columns.str.strip()

        cur = conn.cursor()

        imported = 0
        skipped = 0
        errors = []

        for index, row in df.iterrows():
            try:
                trade_name = str(row[trade_column]).strip() if pd.notna(row[trade_column]) else None
                generic_name = str(row[generic_column]).strip() if pd.notna(row[generic_column]) else None

                if not trade_name or not generic_name:
                    skipped += 1
                    continue

                # Handle multiple generic names (keep them as-is with semicolons)
                # This preserves the format: "dexamethasone; neomycin; polymyxin B"

                # Get medication type if column is mapped
                medication_type = 'Both'  # Default
                if type_column and type_column in row:
                    type_value = str(row[type_column]).strip() if pd.notna(row[type_column]) else None
                    if type_value:
                        # Try to map to our types
                        type_lower = type_value.lower()
                        if 'ocular' in type_lower or 'eye' in type_lower or 'ophthalm' in type_lower:
                            medication_type = 'Ocular'
                        elif 'systemic' in type_lower or 'oral' in type_lower or 'general' in type_lower:
                            medication_type = 'Systemic'
                        else:
                            medication_type = 'Both'

                # Check if medication already exists (by trade name)
                cur.execute('SELECT id FROM medications WHERE trade_name = %s', (trade_name,))
                existing = cur.fetchone()

                if existing:
                    # Update existing medication
                    cur.execute('''
                        UPDATE medications 
                        SET generic_name = %s, medication_type = %s, active = TRUE, updated_at = CURRENT_TIMESTAMP
                        WHERE trade_name = %s
                    ''', (generic_name, medication_type, trade_name))
                else:
                    # Insert new medication
                    cur.execute('''
                        INSERT INTO medications (trade_name, generic_name, medication_type, active)
                        VALUES (%s, %s, %s, TRUE)
                    ''', (trade_name, generic_name, medication_type))

                imported += 1

            except Exception as e:
                errors.append(f"Row {index + 2}: {str(e)}")
                skipped += 1

        conn.commit()
        cur.close()
        conn.close()

        return jsonify({
            'success': True,
            'imported': imported,
            'skipped': skipped,
            'errors': errors[:10] if errors else []  # Return first 10 errors
        })

    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({'error': f'Import failed: {str(e)}'}), 500


@app.route('/api/medications-export')
@admin_required
def medications_export():
    """Export medications to Excel"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection error'}), 500

    try:
        import pandas as pd
        from io import BytesIO
        from datetime import datetime

        # Get all medications
        query = '''
            SELECT 
                trade_name as "Trade Name",
                generic_name as "Generic Name",
                medication_type as "Type",
                CASE WHEN active THEN 'Active' ELSE 'Inactive' END as "Status"
            FROM medications
            ORDER BY trade_name
        '''

        df = pd.read_sql_query(query, conn)
        conn.close()

        # Create Excel file
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Medications', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets['Medications']
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        output.seek(0)

        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename=medications_{datetime.now().strftime("%Y%m%d")}.xlsx'
            }
        )

    except Exception as e:
        if conn:
            conn.close()
        return jsonify({'error': f'Export failed: {str(e)}'}), 500


# Application Initialization and Startup

if __name__ == '__main__':
    print("\n" + "=" * 60)
    print("Starting Flask development server on http://0.0.0.0:5000")
    print("=" * 60 + "\n")

    app.run(host='0.0.0.0', port=5000, debug=False)
